import time
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.conf import settings
import requests
from .models import Beneficiary, APIToken
from datetime import datetime
from django.core.paginator import Paginator
from django.http import JsonResponse,HttpResponse, HttpResponseForbidden, Http404
from django.utils.timezone import localtime
from .forms import APITokenForm
from datetime import datetime, date
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy, reverse
from .forms import StyledPasswordChangeForm
import json
import logging
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.db.models.functions import Lower
from audit.models import AuditLog 
from audit.views import log_user_action
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt  # only if using non-ajax or external POSTs — prefer csrf
from django.db.models.functions import TruncMonth
# wherever you create beneficiary from get_val(...)
from .households import assign_households
from django.db.models import Case, When, Value, CharField, F
from django.db.models.functions import Lower
from openpyxl import load_workbook  # used to read Excel files
import tempfile


## ----------------Permiossion Denied, Role required ----------------------

from django.core.exceptions import PermissionDenied

def role_required(*roles):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)
            if request.user.groups.filter(name__in=roles).exists():
                return view_func(request, *args, **kwargs)
            raise PermissionDenied
        return wrapper
    return decorator

def log_user_action(request, action, model_name=None, record_id=None, changes=None):
    user = request.user if request.user.is_authenticated else None
    AuditLog.objects.create(
        user=user,
        action=action,
        model_name=model_name,
        record_id=record_id,
        changes=changes or {},
        timestamp=timezone.now()
    )



@login_required
def home(request):
    # Gender distribution
    # Aggregate gender counts

    # Get distinct sectors for the filter dropdown
    sector_list = (
        Beneficiary.objects
        .exclude(Sector__isnull=True)
        .exclude(Sector='')
        .values_list('Sector', flat=True)
        .distinct()
        .order_by('Sector')
    )

    selected_sector = request.GET.get('sector', '')

    base_qs = Beneficiary.objects.all()

    if selected_sector:
        base_qs = base_qs.filter(Sector=selected_sector)


    gender_data = base_qs.values('Gender').annotate(total=Count('id'))

    # Initialize counters
    gender_counts = {'Male': 0, 'Female': 0, 'Unknown': 0}

    # Normalize and sum
    for g in gender_data:
        gender = (g['Gender'] or '').strip().lower()
        if gender in ['m', 'male']:
            gender_counts['Male'] += g['total']
        elif gender in ['f', 'female']:
            gender_counts['Female'] += g['total']
        else:
            gender_counts['Unknown'] += g['total']

    # Prepare for JS
    gender_labels = ['Male', 'Female', 'Unknown']
    gender_values = [gender_counts['Male'], gender_counts['Female'], gender_counts['Unknown']]

    
    # Age groups (example buckets)
    age_buckets = {
        '0-17': base_qs.filter(Age__lte=17).count(),
        '18-35': base_qs.filter(Age__gte=18, Age__lte=35).count(),
        '36-60': base_qs.filter(Age__gte=36, Age__lte=60).count(),
        '60+': base_qs.filter(Age__gte=61).count(),
    }

    # Sector
    sector_data = base_qs.values('IP_Name').annotate(total=Count('id'))
    sector_labels = [s['IP_Name'] or 'Unknown' for s in sector_data]
    sector_values = [s['total'] for s in sector_data]

    
    # Aggregate submissions by month and IP_Name
    submissions = (
        base_qs
        .annotate(month=TruncMonth('Submission_Time'))
        .values('month', 'IP_Name')
        .annotate(total=Count('id'))
        .order_by('month', 'IP_Name')
    )

    # All IP names in the dataset
    ip_names = sorted({s['IP_Name'] or "Unknown" for s in submissions})

    # Sorted months
    submission_months = sorted({s['month'] for s in submissions if s['month']})
    submission_labels = [m.strftime("%b %Y") for m in submission_months]

    # Map month objects to their index
    month_index_map = {m: i for i, m in enumerate(submission_months)}

    # Initialize datasets
    datasets = {ip: [0] * len(submission_labels) for ip in ip_names}

    # Fill datasets
    for s in submissions:
        if s['month'] is None:
            continue
        ip = s['IP_Name'] or "Unknown"
        idx = month_index_map[s['month']]
        datasets[ip][idx] = s['total']

    # Chart.js datasets
    submission_datasets = [{"label": ip, "data": datasets[ip]} for ip in ip_names]


    # Governorates
    
    gov_data = (
        base_qs
        .annotate(
            gov_raw=Case(
                When(Governorate__isnull=True, then=Value('Unknown')),
                When(Governorate='', then=Value('Unknown')),
                default=F('Governorate'),
                output_field=CharField(),
            ),
            gov_cleaned=Lower(
                Case(
                    When(Governorate__isnull=True, then=Value('Unknown')),
                    When(Governorate='', then=Value('Unknown')),
                    default=F('Governorate'),
                    output_field=CharField(),
                )
            )
        )
        .values('gov_cleaned')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )


    gov_labels = [
        g['gov_cleaned'].capitalize() if g['gov_cleaned'] != "unknown" else "Unknown"
        for g in gov_data
    ]

    gov_values = [g['total'] for g in gov_data]



    last_sync_record = Beneficiary.all_objects.order_by('-created_at').first()

    

    context = {
        'sector_list': sector_list,
        'selected_sector': selected_sector,
        "gender_labels": gender_labels,
        "gender_data": gender_values,
        "age_labels": list(age_buckets.keys()),
        "age_data": list(age_buckets.values()),
        "sector_labels": sector_labels,
        "sector_data": sector_values,
        "submission_labels": submission_labels,
        "submission_datasets": submission_datasets,
        "gov_labels": gov_labels,
        "gov_data": gov_values,
        
        "last_sync_record": Beneficiary.all_objects.order_by('-created_at').first(),
        "last_sync_time" : localtime(last_sync_record.created_at).strftime("%Y-%m-%d %H:%M") if last_sync_record else "1900-01-01 00:00",
    
    }
    return render(request, "myproject/home.html", context)

     
    
    return render(request, 'myproject/home.html', {'last_sync_time': last_sync_time})


def get_item(dictionary, key):
    if isinstance(dictionary, dict):
        return dictionary.get(key, "")
    return ""


import os, threading, time

# global variable for progress
progress_status = {"stage": "Not started", "progress": 0}

@login_required
@role_required('Manager', 'Admin')
def sync_data(request):
    global progress_status
    
    if request.method != 'POST':
        return JsonResponse({"status": "error", "message": "Invalid request method. Only POST allowed.", "extra_tags":"auto"})

    # Reset before new run
    progress_status = {"stage": "Starting sync ...", "progress": 1, "completed": False}

    def run_sync():
        global progress_status
        tokens = APIToken.objects.all()
        if not tokens.exists():
            progress_status = {"stage": "❌ No API tokens available.", "progress": 0, "completed": True}
            return

        total_fetched = 0
        new_count = 0
        duplicate_count = 0
        invalid_date_count = 0

        # iterate tokens but do NOT let one token failure abort everything
        for token in tokens:
            token_id = getattr(token, "form_id", str(token.pk))
            page = 1
            page_size = 10000

            last_sync_record = Beneficiary.all_objects.order_by('-created_at').first()
            last_sync_time = (
                localtime(last_sync_record.created_at).strftime("%Y-%m-%d %H:%M")
                if last_sync_record
                else "1900-01-01 00:00"
            )

            # token-level try so we continue to next token on failure
            try:
                while True:
                    url = f"https://data.inform.unicef.org/api/v1/data/{token.form_id}.json?page={page}&page_size={page_size}"
                    headers = {"Authorization": f"Token {token.token}", "Content-Type": "application/json"}

                    try:
                        response = requests.get(url, headers=headers, timeout=30)
                    except requests.RequestException as e:
                        # network/timeout error — report and continue to next token
                        progress_status = {
                            "stage": f"⚠️ Network error for token {token_id}: {str(e)}",
                            "progress": 0,
                            "completed": False,
                        }
                        break  # stop pages for this token, move to next token

                    if response.status_code == 401 or response.status_code == 403:
                        # unauthorized for this token — skip token
                        progress_status = {
                            "stage": f"⚠️ Auth failed for token {token_id} (HTTP {response.status_code}) — skipping token",
                            "progress": 0,
                            "completed": False,
                        }
                        break

                    if response.status_code != 200:
                        # other non-200 — log and stop paging this token
                        progress_status = {
                            "stage": f"⚠️ Failed fetching token {token_id} page {page} (HTTP {response.status_code})",
                            "progress": 0,
                            "completed": False,
                        }
                        break

                    # parse JSON safely
                    try:
                        data = response.json()
                        # ✅ Stop when there’s no data or an empty page
                        if not data or (isinstance(data, dict) and not data.get("data") and not data.get("results")):
                            break
                        # If API wraps data under "data"
                        if isinstance(data, dict) and "data" in data:
                            data = data["data"]

                        if len(data) == 0:
                            break

                        total_fetched += len(data)
                    except ValueError:
                        progress_status = {
                            "stage": f"⚠️ Invalid JSON for token {token_id} page {page} — skipping token",
                            "progress": 0,
                            "completed": False,
                        }
                        break

                    if not data:
                        # no more pages
                        break

                    # increment totals for reporting
                    #total_fetched += len(data)

                    for idx, item in enumerate(data, start=1):
                        def get_val(key):
                            val = item.get(key)
                            if isinstance(val, (tuple, list)):
                                val = val[0] if val else None
                            if isinstance(val, str):
                                val = val.strip()
                                if val.lower() in ["", "null", "n/a", "none"]:
                                    return None
                            return val

                        # progress update (per-record)
                        progress_status = {
                            "stage": f"IP: {get_val('IP_Name')} | Form: {token_id} | Page: {page} | Record:({idx}/{len(data)})",
                            # show progress up to 98 while running; final will be 100 when done
                            "progress": min(98, int((idx / max(len(data), 1)) * 100)),
                            "new_count": new_count,
                            "duplicate_count": duplicate_count,
                            "total_fetched": total_fetched,
                            "completed": False,
                        }
                        InForm_ID = get_val("_id")
                        

                        #last_synced = get_val('_submission_time')
                        # Try to find an existing beneficiary with this InForm_ID
                        
                        #new_rec = last_synced > last_sync_time
                        #print(f"New record check: {new_rec} for submission time {last_synced} vs last sync {last_sync_time}")
                        #if new_rec:
                        if Beneficiary.all_objects.filter(record_id=InForm_ID).exists():
                            continue  # skip duplicates
                            # Create new record only if not found or InForm_ID is empty
                        Beneficiary.all_objects.update_or_create(
                            record_id=get_val("_id"),
                            defaults={
                                "record_id": get_val("_id"),
                                "InForm_ID": get_val("_id"),
                                "InstanceID": f"uuid:{get_val('_uuid')}",
                                "IP_Name": get_val("IP_Name"),
                                "Sector": get_val("Sector"),
                                "Indicator": get_val("Indicator"),
                                "Date": get_val("Date"),
                                "Name": get_val("Name"),
                                "ID_Number": get_val("ID_Number"),
                                "Parent_ID": get_val("Parent_ID"),
                                "Spouse_ID": get_val("Spouse_ID"),
                                "Phone_Number": get_val("Phone_Number"),
                                "Date_of_Birth": get_val("Date_of_Birth"),
                                "Age": get_val("Age"),
                                "Gender": get_val("Gender"),
                                "Governorate": get_val("Governorate"),
                                "Municipality": get_val("Municipality"),
                                "Neighborhood": get_val("Neighborhood"),
                                "Site_Name": get_val("Site_Name"),
                                "Disability_Status": get_val("Disability_Status"),
                                "Submission_Time": get_val("_submission_time"),
                                "Deleted": get_val("Deleted"),
                                "deleted_at": get_val("deleted_at"),
                                "undeleted_at": get_val("undeleted_at"),
                                
                            }
                        )
                        new_count += 1

                    # next page for this token
                    # ✅ Move to next page only if this page was full
                    if len(data) < page_size:
                        # means this was the last page
                        break

                    page += 1

            except Exception as e_token:
                # catch unexpected exceptions for this token and continue to the next
                progress_status = {
                    "stage": f"❌ Unexpected error while processing token {token_id}: {str(e_token)}",
                    "progress": 0,
                    "completed": False,
                }
                # continue to next token
                continue

        # After all tokens processed, assign households and produce final summary
        try:
            changes = assign_households()
            messages.success(request, f"Household assignment updated for {changes} records.", extra_tags="auto")
        except Exception as e_house:
            changes = 0
            # log or include house error in the status
            progress_status = {
                "stage": f"⚠️ Household assignment error: {str(e_house)}",
                "progress": 100,
                "completed": True,
            }
        messages.success(request, f"Household assignment updated for {changes} records.", extra_tags="auto")

        progress_status = {
            "stage": "✅ Sync completed",
            "progress": 100,
            "new_count": new_count,
            "duplicate_count": duplicate_count,
            "invalid_date_count": invalid_date_count,
            "total_fetched": total_fetched,
            "household_changes": changes,
            "completed": True,
            "extra_tags":"auto"
        }

    thread = threading.Thread(target=run_sync, daemon=True)
    thread.start()

    return JsonResponse({"status": "started"})
    
    

    


def get_progress(request):
    global progress_status
    return JsonResponse(progress_status)



def inform_data_view(request):

    # Admins/Managers see everything
    if request.user.is_superuser or request.user.groups.filter(name__in=["Manager", "Admin"]).exists():
        submissions = Beneficiary.objects.all()

    # Normal staff only see their own records
    else:
        submissions = Beneficiary.objects.filter(created_by=request.user)

    # --- Sorting ---
    sort_field = request.GET.get('sort', '-created_at')  # default sort
    # Ensure the sort field is valid
    valid_fields = [f.name for f in Beneficiary._meta.fields]
    if sort_field.lstrip('-') in valid_fields:
        submissions = submissions.order_by(sort_field)
    else:
        submissions = submissions.order_by('-created_at')

    # --- Filtering ---
    filter_fields = ['IP_Name','Sector','Indicator','Date','Name','ID_Number','Parent_ID','Spouse_ID','Phone_Number',
                     'Date_of_Birth','Age','Gender','Governorate','Household_ID']
    
    from django.db.models import Q

    operator = request.GET.get('operator', 'AND').upper()

    # Build Q objects for each filter field
    q_objects = []
    for field in filter_fields:  # filter_fields should be list of (field_name, label)
        value = request.GET.get(field[0], '').strip()
        if value:
            q_objects.append(Q(**{f"{field[0]}__icontains": value}))

    
    # Apply AND / OR
    qs = Beneficiary.objects.all()
    if operator == 'AND':
        for q in q_objects:
            qs = qs.filter(q)
    #elif operator == 'OR' and q_objects:
    #    combined_q = Q()
    #    for q in q_objects:
    #        combined_q |= q
    #    qs = qs.filter(combined_q)

    for field in filter_fields:
        value = request.GET.get(field)
        if value:
            submissions = submissions.filter(**{f"{field}__icontains": value})
    

    cnt = Beneficiary.objects.count()  # total in DB

    # --- Pagination ---
    paginator = Paginator(submissions, 18)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Fields for filter inputs (friendly labels)
    fields = [
        ('IP_Name','IP Name'),
        ('Sector','Sector'),
        ('Indicator','Indicator'),
        ('Date','Date'),
        ('Name','Name'),
        ('ID_Number','ID Number'),
        ('Parent_ID','Parent_ID'),
        ('Spouse_ID','Spouse_ID'),
        ('Phone_Number','Phone Number'),
        ('Date_of_Birth','Date of Birth'),
        ('Age','Age'),
        ('Gender','Gender'),
        ('Governorate','Governorate'),
        ('Household_ID','Household_ID'),
    ]

    context = {
        'submissions': page_obj,
        'filter_fields': fields,
        'request': request,  # for pagination links
        'total_records': submissions.count(),  # total after filtering
        'total_all_records': cnt,  # total in DB
        'current_sort': sort_field,  # current sorting
    }

    # AJAX request returns only table
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'myproject/inform_table.html', context)

    return render(request, 'myproject/inform_data.html', context)




@login_required
@role_required('Manager', 'Admin')
def manage_tokens(request):
    if request.method == 'POST':
        form = APITokenForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Token saved successfully.", extra_tags="auto")
            return redirect('manage_tokens')
        else:
            messages.error(request, "❌ Please correct the errors below.", extra_tags="auto")
    else:
        form = APITokenForm()

    tokens = APIToken.objects.all().order_by('-id')  # Optional: newest on top
    return render(request, 'myproject/manage_tokens.html', {
        'form': form,
        'tokens': tokens
    })

@login_required
@role_required('Manager', 'Admin')
def delete_token(request, pk):
    token = get_object_or_404(APIToken, pk=pk)
    token.delete()
    messages.success(request, "API token deleted successfully.", extra_tags="auto")
    return redirect('manage_tokens')  # Redirect back to the manage tokens page







def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, "You have successfully logged in.", extra_tags="auto")
            # NOW user is a real User, safe to log
            AuditLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action="LOGIN",
                model_name="User",
                record_id=str(request.user.id) if request.user.is_authenticated else None,
                changes={},  # or any relevant info
                description="User Logged In",
                timestamp=timezone.now()
            )

            return redirect('home')  # Redirect to your desired page
        else:
            messages.error(request, "Invalid username or password.", extra_tags="auto")
    else:
        form = AuthenticationForm()

    
    context = {
        'form': form,
    }
    return render(request, 'myproject/login.html', context)

def user_logout(request):
    
    AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action="LOGOUT",
            model_name="User",
            record_id=str(request.user.id) if request.user.is_authenticated else None,
            changes={},  # or any relevant info
            description="User Logged Out",
            timestamp=timezone.now()
        )

    logout(request)
    
    messages.success(request, "You have been logged out.", extra_tags="auto")
    return redirect("login")  # Replace with your login URL name



def user_register(request):
    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if password == confirm_password:
            if User.objects.filter(username=username).exists():
                messages.error(request, "Username already exists.", extra_tags="auto")
            elif User.objects.filter(email=email).exists():
                messages.error(request, "Email is already registered.", extra_tags="auto")
            else:
                user = User.objects.create_user(username=username, email=email, first_name=first_name, last_name=last_name, password=password)
                user.is_active = True  # 🔑 make sure the account is active
                user.save()
                print(user.password)
                messages.success(request, "Account created successfully! Please log in.", extra_tags="auto")
                return redirect('login')
        else:
            messages.error(request, "Passwords do not match.", extra_tags="auto")
    return render(request, 'myproject/register.html')


@login_required
def profile(request):
    if request.method == "POST":
        user = request.user
        user.first_name = request.POST.get("first_name")
        user.last_name = request.POST.get("last_name")
        user.email = request.POST.get("email")
        user.save()
        messages.success(request, "User profile updated successfully!", extra_tags="auto")
        return redirect("home")
    return render(request, "myproject/profile.html")

@login_required
def password_change(request):
    if request.method == "POST":
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Keep user logged in after password change
            update_session_auth_hash(request, user)
            messages.success(request, "✅ Your password was successfully updated!", extra_tags="auto")
            return redirect("home")  # redirect to profile or dashboard
        else:
            messages.error(request, "❌ Please correct the errors below.", extra_tags="auto")
    else:
        form = PasswordChangeForm(request.user)

    return render(request, "myproject/password_change.html", {"form": form})


class CustomPasswordChangeView(PasswordChangeView):
    form_class = StyledPasswordChangeForm
    template_name = "myproject/change_password.html"
    success_url = reverse_lazy('home')


# ------- ID_Number Duplication Section -----------------------------------------------------------

logger = logging.getLogger(__name__)
@login_required
def find_duplicates(request):
    """
    Page that finds duplicates in Beneficiary table grouped by id_number,
    and displays each duplicate set in its own table.
    """
    # Adjust `id_number` field name if your model uses different casing.
    # Group by id_number and keep groups with count>1

    # Admins/Managers see everything
    if request.user.is_superuser or request.user.groups.filter(name__in=["Manager", "Admin"]).exists():
        submissions = Beneficiary.objects.all()

    # Normal staff only see their own records
    else:
        submissions = Beneficiary.objects.filter(created_by=request.user)


    total_records = 0
    duplicate_groups = (
        submissions
        .values("ID_Number")
        .annotate(count=Count("ID_Number"))
        .filter(count__gt=1)
        .order_by("-count")
    )

    # Build a list of groups where each group contains the actual records
    groups = []
    for g in duplicate_groups:
        ID_Number = g["ID_Number"]
        records = list(Beneficiary.objects.filter(ID_Number=ID_Number).order_by('-created_at'))
        groups.append({
            "ID_Number": ID_Number,
            "count": g["count"],
            "records": records,
        })
        total_records += len(records)

    context = {
        "groups": groups,
        "total_duplicates": len(groups),
        "total_records": total_records,
    }
    return render(request, "myproject/duplicates.html", context)


@require_POST
@login_required
def delete_inform(request, pk):
    """
    Deletes a single Beneficiary record (pk).
    Expects POST data:
      - delete_source: 'true' or 'false' (whether to attempt removal from InForm API)
    Returns JSON (success/error) — used by AJAX in the template.
    """
    if not request.user.is_authenticated:
        return HttpResponseForbidden("Authentication required.")

    delete_source = request.POST.get("delete_source", "false").lower() == "true"

    # Get target record
    obj = get_object_or_404(Beneficiary, pk=pk)
    inform_record_id = getattr(obj, "InForm_ID", None)  # adjust if field differs
    ip_name = getattr(obj, "IP_Name", None)

    if delete_source:
        if not inform_record_id:
            return JsonResponse({
                "success": False,
                "error": "Record has no linked InForm record id (InForm_ID). Cannot delete from InForm."
            }, status=400)

        # ---------- Prepare InForm Payload ----------
        import json
        import uuid
        from django.utils import timezone
        import requests
        from django.contrib import messages

        form_id_string = get_form_id_string(obj.IP_Name)
        inform_api_url = "https://data.inform.unicef.org/unicefstateofpalestine/submission"
        api_token = "ef5a72dad573916ed3a2289e51ca4f1b8ced5c88"  # replace with real one
        new_instance_id = f"uuid:{str(uuid.uuid4())}"
        obj.Deleted = True
        obj.deleted_at = timezone.now().strftime("%Y-%m-%dT%H:%M:%S")

        payload = {
            "id": form_id_string,
            "submission": {
                "IP_Name": obj.IP_Name,
                "Sector": obj.Sector,
                "Indicator": obj.Indicator,
                "Date": obj.Date,
                "Name": obj.Name,
                "ID_Number": obj.ID_Number,
                "Parent_ID": obj.Parent_ID,
                "Spouse_ID": obj.Spouse_ID,
                "Phone_Number": obj.Phone_Number,
                "Date_of_Birth": obj.Date_of_Birth,
                "Age": obj.Age,
                "Gender": obj.Gender,
                "Governorate": obj.Governorate,
                "Municipality": obj.Municipality,
                "Neighborhood": obj.Neighborhood,
                "Site_Name": obj.Site_Name,
                "Disability_Status": obj.Disability_Status,
                "Deleted": obj.Deleted,
                "deleted_at": obj.deleted_at,
                "undeleted_at": obj.undeleted_at.strftime("%Y-%m-%dT%H:%M:%S") if obj.undeleted_at else None,
                "meta": {
                    "instanceID": new_instance_id,
                    "deprecatedID": str(obj.InstanceID) if obj.InstanceID else None,
                },
            },
        }

        headers = {
            "Authorization": f"Token {api_token}",
            "Content-Type": "application/json",
        }

        # ---------- Send to InForm ----------
        try:
            # ✅ Convert non-serializable types automatically
            json_payload = json.dumps(payload, default=str)

            response = requests.post(inform_api_url, headers=headers, data=json_payload, timeout=20)

            if response.status_code in [200, 201]:
                obj.InstanceID = new_instance_id
                obj.save()
                messages.success(request, "✅ Record has been deleted successfully and synced to InForm.", extra_tags="auto")
            else:
                messages.warning(
                    request,
                    f"⚠️ Local update saved but InForm sync failed ({response.status_code}): {response.text}",
                    extra_tags="auto"
                )
        except Exception as e:
            messages.error(request, f"❌ Local update saved but failed to connect to InForm: {str(e)}", extra_tags="auto")


    # Delete the local DB record
    obj.soft_delete()

    return JsonResponse({"success": True, "deleted_pk": pk})

@require_POST
@login_required
def delete_duplicate(request, pk):
    """
    Deletes a single Beneficiary record (pk).
    Expects POST data:
      - delete_source: 'true' or 'false' (whether to attempt removal from InForm API)
    Returns JSON (success/error) — used by AJAX in the template.
    """
    if not request.user.is_authenticated:
        return HttpResponseForbidden("Authentication required.")

    delete_source = request.POST.get("delete_source", "false").lower() == "true"

    # Get target record
    obj = get_object_or_404(Beneficiary, pk=pk)
    inform_record_id = getattr(obj, "InForm_ID", None)  # adjust if field differs
    ip_name = getattr(obj, "IP_Name", None)

        # ---------- Prepare InForm Payload ----------
    if delete_source:
        form_id_string = get_form_id_string(obj.IP_Name)
        inform_api_url = "https://data.inform.unicef.org/unicefstateofpalestine/submission"
        api_token = "ef5a72dad573916ed3a2289e51ca4f1b8ced5c88"  # replace with real one
        new_instance_id = f"uuid:{str(uuid.uuid4())}"
        obj.Deleted = True
        obj.deleted_at = timezone.now().strftime("%Y-%m-%dT%H:%M:%S")

        payload = {
            "id": form_id_string,
            "submission": {
                "IP_Name": obj.IP_Name,
                "Sector": obj.Sector,
                "Indicator": obj.Indicator,
                "Date": obj.Date,
                "Name": obj.Name,
                "ID_Number": obj.ID_Number,
                "Parent_ID": obj.Parent_ID,
                "Spouse_ID": obj.Spouse_ID,
                "Phone_Number": obj.Phone_Number,
                "Date_of_Birth": obj.Date_of_Birth,
                "Age": obj.Age,
                "Gender": obj.Gender,
                "Governorate": obj.Governorate,
                "Municipality": obj.Municipality,
                "Neighborhood": obj.Neighborhood,
                "Site_Name": obj.Site_Name,
                "Disability_Status": obj.Disability_Status,
                "Deleted": obj.Deleted,
                "deleted_at": obj.deleted_at,
                "undeleted_at": obj.undeleted_at.strftime("%Y-%m-%dT%H:%M:%S") if obj.undeleted_at else None,
                "meta": {
                    "instanceID": new_instance_id,
                    "deprecatedID": obj.InstanceID,
                },
            },
        }

        headers = {
            "Authorization": f"Token {api_token}",
            "Content-Type": "application/json",
        }

        # ---------- Send to InForm ----------
        try:
            # ✅ Convert non-serializable types automatically
            json_payload = json.dumps(payload, default=str)

            response = requests.post(inform_api_url, headers=headers, data=json_payload, timeout=20)

            if response.status_code in [200, 201]:
                obj.InstanceID = new_instance_id
                obj.save()
                messages.success(request, "✅ Record has been deleted successfully and synced to InForm.", extra_tags="auto")
            else:
                messages.warning(
                    request,
                    f"⚠️ Local update saved but InForm sync failed ({response.status_code}): {response.text}",
                    extra_tags="auto"
                )
        except Exception as e:
            messages.error(request, f"❌ Local update saved but failed to connect to InForm: {str(e)}", extra_tags="auto")

    # Delete the local DB record
    obj.soft_delete()
    messages.success(request, "✅ Record has been deleted successfully from local database.", extra_tags="auto")

    return JsonResponse({"success": True, "deleted_pk": pk})


# ---------- Beneficiary Details page ---------------------------------------------------------------

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
import requests, uuid
from .models import Beneficiary

def safe_value(val):
    """Convert empty strings to None for API submission."""
    return val if val not in ["", None] else None


def get_form_id_string(ip_name):
    """Return correct Form_ID_String based on IP_Name."""
    mapping = {
        "AEI": "Beneficiary_Database_Form_Template",
        "UNRWA": "Beneficiary_Database_Form_Template_-_UNRWA",
        "WeWorld": "Beneficiary_Database_Form_Template_-_We_World",
        "MedGlobal": "azPSg7utHm6x4kNyfpdgCj",
        "RRM": "Beneficiary_Database_Form_Template_-_RRM",
        "NECC": "Beneficiary_Database_Form_Template_-_NECC",
        "Taawon": "Beneficiary_Database_Form_Template_-_Taawon",
    }
    return mapping.get(ip_name, "Beneficiary_Database_Form_Template_-_Default")


def beneficiary_details(request, pk):
    record = get_object_or_404(Beneficiary, pk=pk)

    if request.method == "POST":
        # ---------- UPDATE LOCAL FIELDS ----------
        old_instance_id = record.InstanceID
        new_instance_id = f"uuid:{uuid.uuid4()}"

        record.IP_Name = safe_value(request.POST.get("IP_Name"))
        record.Sector = request.POST.get("Sector")
        record.Indicator = request.POST.get("Indicator")
        record.Date = safe_value(request.POST.get("Date"))
        record.Name = request.POST.get("Name")
        record.ID_Number = request.POST.get("ID_Number")
        record.Parent_ID = request.POST.get("Parent_ID")
        record.Spouse_ID = request.POST.get("Spouse_ID")
        record.Phone_Number = request.POST.get("Phone_Number")
        record.Date_of_Birth = safe_value(request.POST.get("Date_of_Birth"))
        record.Age = request.POST.get("Age")
        record.Gender = request.POST.get("Gender")
        record.Governorate = request.POST.get("Governorate")
        record.Municipality = request.POST.get("Municipality")
        record.Neighborhood = request.POST.get("Neighborhood")
        record.Site_Name = request.POST.get("Site_Name")
        record.Disability_Status = request.POST.get("Disability_Status")

        record.save()

        # ---------- Update Households ----------
        changes = assign_households()
        #messages.success(request, f"Local record updated and household assignment updated for {changes} records.", extra_tags="auto")

        # ---------- Prepare InForm Payload ----------
        form_id_string = get_form_id_string(record.IP_Name)
        inform_api_url = "https://data.inform.unicef.org/unicefstateofpalestine/submission"
        api_token = "ef5a72dad573916ed3a2289e51ca4f1b8ced5c88"  # replace with real one

        payload = {
            "id": form_id_string,
            "submission": {
                "start": None,
                "end": None,
                "Name": safe_value(record.Name),
                "ID_Number": safe_value(record.ID_Number),
                "Parent_ID": safe_value(record.Parent_ID),
                "Spouse_ID": safe_value(record.Spouse_ID),
                "Phone_Number": safe_value(record.Phone_Number),
                "Date_of_Birth": safe_value(record.Date_of_Birth),
                "Age": safe_value(record.Age),
                "Gender": safe_value(record.Gender),
                "Governorate": safe_value(record.Governorate),
                "Municipality": safe_value(record.Municipality),
                "Neighborhood": safe_value(record.Neighborhood),
                "Site_Name": safe_value(record.Site_Name),
                "Disability_Status": safe_value(record.Disability_Status),
                "Sector": safe_value(record.Sector),
                "IP_Name": safe_value(record.IP_Name),
                "Indicator": safe_value(record.Indicator),
                "Date": safe_value(record.Date),
                "meta": {
                    "instanceID": new_instance_id,
                    "deprecatedID": old_instance_id,
                },
            },
        }

        headers = {
            "Authorization": f"Token {api_token}",
            "Content-Type": "application/json",
        }

        # ---------- Send to InForm ----------
        try:
            response = requests.post(inform_api_url, headers=headers, json=payload, timeout=20)
            if response.status_code in [200, 201]:
                record.InstanceID = new_instance_id
                record.save()
                messages.success(request, f"✅ Record updated, synced to InForm and households updated for {changes} records..", extra_tags="auto")
            else:
                messages.warning(
                    request,
                    f"⚠️ Local update saved but InForm sync failed ({response.status_code}): {response.text}",
                    extra_tags="auto"
                )
        except Exception as e:
            messages.error(request, f"❌ Local update saved but failed to connect to InForm: {str(e)}", extra_tags="auto")

        return redirect("beneficiary_details", pk=record.pk)
    
    
    # Prepare filter options
    ip_list = APIToken.objects.values_list('IP', flat=True).distinct().order_by('IP')
    sector_list = APIToken.objects.values_list('section', flat=True).distinct().order_by('section')

    context = {
        "record": record,
        'ip_list': ip_list,
        'sector_list': sector_list,
        # include any other variables you already had in context
    }
    # ---------- GET request: display details page ----------
    return render(request, "myproject/beneficiary_details.html", context)



# ---------- Household Details page ---------------------------------------------------------------


@login_required
def household_details(request, household_id):
    # Get all members in this household
    members = Beneficiary.objects.filter(Household_ID=household_id)

    if not members.exists():
        raise Http404("Household not found")

    # Identify potential head (Parent_ID is empty, usually the head)
    head = next((m for m in members if not m.Parent_ID), members.first())

    # Build relationship mapping
    relationships = []
    for member in members:
        relation = "Head" if member == head else "Spouse" if member.Spouse_ID == head.ID_Number else "Child"
        relationships.append({
            "relation": relation,
            "member": member,
        })

    context = {
        "household_id": household_id,
        "head": head,
        "relationships": relationships,
        "total_members": members.count(),
    }
    return render(request, "myproject/household_details.html", context)

# ---------- Deleted Beneficiaries page ---------------------------------------------------------------

@login_required
def deleted_beneficiaries(request):

    # Admins/Managers see everything
    if request.user.is_superuser or request.user.groups.filter(name__in=["Manager", "Admin"]).exists():
        submissions = Beneficiary.deleted_objects.all()

    # Normal staff only see their own records
    else:
        submissions = Beneficiary.deleted_objects.filter(created_by=request.user)

    deleted_records = submissions.order_by('-deleted_at')
    
    if request.method == "POST":
        pk = request.POST.get("pk")
        if pk:
            record = get_object_or_404(Beneficiary.deleted_objects, pk=pk)
            record.Deleted = False
            record.undeleted_at = timezone.now().strftime("%Y-%m-%dT%H:%M:%S")
            
            # ---------- Prepare InForm Payload ----------
            form_id_string = get_form_id_string(record.IP_Name)
            inform_api_url = "https://data.inform.unicef.org/unicefstateofpalestine/submission"
            api_token = "ef5a72dad573916ed3a2289e51ca4f1b8ced5c88"  # replace with real one
            new_instance_id = f"uuid:{str(uuid.uuid4())}"

            payload = {
                "id": form_id_string,
                "submission": {
                    "IP_Name": record.IP_Name,
                    "Sector": record.Sector,
                    "Indicator": record.Indicator,
                    "Date": record.Date,
                    "Name": record.Name,
                    "ID_Number": record.ID_Number,
                    "Parent_ID": record.Parent_ID,
                    "Spouse_ID": record.Spouse_ID,
                    "Phone_Number": record.Phone_Number,
                    "Date_of_Birth": record.Date_of_Birth,
                    "Age": record.Age,
                    "Gender": record.Gender,
                    "Governorate": record.Governorate,
                    "Municipality": record.Municipality,
                    "Neighborhood": record.Neighborhood,
                    "Site_Name": record.Site_Name,
                    "Disability_Status": record.Disability_Status,
                    "Deleted": record.Deleted,
                    "deleted_at": record.deleted_at.strftime("%Y-%m-%dT%H:%M:%S") if record.undeleted_at else None,
                    "undeleted_at": record.undeleted_at,
                    "meta": {
                        "instanceID": new_instance_id,
                        "deprecatedID": record.InstanceID,
                    },
                },
            }

            headers = {
                "Authorization": f"Token {api_token}",
                "Content-Type": "application/json",
            }

            # ---------- Send to InForm ----------
            try:
                json_payload = json.dumps(payload, default=str)
                response = requests.post(inform_api_url, headers=headers, data=json_payload, timeout=20)

                if response.status_code in [200, 201]:
                    record.InstanceID = new_instance_id
                    record.save()
                    messages.success(request, "✅ Record restored successfully and synced to InForm.", extra_tags="auto")
                else:
                    record.save()
                    messages.warning(
                        request,
                        f"⚠️ Record restored locally but InForm sync failed ({response.status_code}): {response.text}",
                        extra_tags="auto"
                    )
            except Exception as e:
                record.save()
                messages.error(request, f"❌ Record restored locally but failed to connect to InForm: {str(e)}", extra_tags="auto")

            return redirect('deleted_beneficiaries')

    return render(request, "myproject/deleted_beneficiaries.html", {"records": deleted_records})



# ---------- Adding Beneficiaries page ---------------------------------------------------------------

import uuid
import requests
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Beneficiary
from .households import assign_households


@login_required
def beneficiary_add(request):
    if request.method == "POST":
        id_number = request.POST.get("ID_Number")
        
        # 🔹 Step 1: Check duplicates locally
        duplicates = Beneficiary.all_objects.filter(ID_Number=id_number)
        if duplicates.exists():
            messages.warning(request, f"⚠️ Duplicate beneficiaries found with ID {id_number}.", extra_tags="auto")
            return render(request, "myproject/duplicate_list.html", {"duplicates": duplicates})

        # 🔹 Step 2: Prepare payload for InForm API
        new_uuid=uuid.uuid4()
        new_instance_id = f"uuid:{new_uuid}"
        form_data = {
            "IP_Name": safe_value(request.POST.get("IP_Name")),
            "Sector": safe_value(request.POST.get("Sector")),
            "Indicator": safe_value(request.POST.get("Indicator")),
            "Date": safe_value(request.POST.get("Date")),
            "Name": safe_value(request.POST.get("Name")),
            "ID_Number": safe_value(request.POST.get("ID_Number")),
            "Parent_ID": safe_value(request.POST.get("Parent_ID")),
            "Spouse_ID": safe_value(request.POST.get("Spouse_ID")),
            "Phone_Number": safe_value(request.POST.get("Phone_Number")),
            "Date_of_Birth": safe_value(request.POST.get("Date_of_Birth")),
            "Age": safe_value(request.POST.get("Age")),
            "Gender": safe_value(request.POST.get("Gender")),
            "Governorate": safe_value(request.POST.get("Governorate")),
            "Municipality": safe_value(request.POST.get("Municipality")),
            "Neighborhood": safe_value(request.POST.get("Neighborhood")),
            "Site_Name": safe_value(request.POST.get("Site_Name")),
            "Disability_Status": safe_value(request.POST.get("Disability_Status")),
            "created_by": request.user,
        }

        # --- Prepare InForm API payload ---
        api_token = "ef5a72dad573916ed3a2289e51ca4f1b8ced5c88"  # your real API token
        inform_api_url = "https://data.inform.unicef.org/unicefstateofpalestine/submission"
        form_id_string = get_form_id_string(form_data["IP_Name"])  # your helper
        # get IP name from form data
        ip_name = form_data.get("IP_Name")
        payload = {
            "id": form_id_string,
            "submission": {
                **form_data,
                "meta": {
                    "instanceID": new_instance_id,
                },
            },
        }
        headers = {
            "Authorization": f"Token {api_token}",
            "Content-Type": "application/json",
        }

        # 🔹 Step 3: Submit to InForm API
        try:
            response = requests.post(inform_api_url, headers=headers, json=payload, timeout=20)

            if response.status_code in [200, 201]:
                
                try:
                    apitoken = APIToken.objects.get(IP=ip_name)
                    form_id = apitoken.form_id
                    token_value = apitoken.token
                    print(f"Form ID: {form_id}, Token: {token_value}, InstanceID: {new_instance_id}")
                    print(f"Posted to InForm")
                except APIToken.DoesNotExist:
                    print(f"No APIToken found for IP: {ip_name}")
                    form_id = None
                    token_value = None

                # 🔹 Step 4: Get the InForm -id
                """
                Fetch one record from InForm API by InstanceID and save/update it locally.
                """
                uuid_str = str(new_uuid) if not isinstance(new_uuid, str) else new_uuid
                query = json.dumps({"_uuid": uuid_str})
                url = f"https://data.inform.unicef.org/api/v1/data/{form_id}.json?query={query}"
                headers = {"Authorization": f"Token {token_value}", "Content-Type": "application/json"}

                try:
                    response = requests.get(url, headers=headers, timeout=30)
                    response.raise_for_status()
                    print("✅ Connected to InForm")
                except requests.RequestException as e:
                    print(f"⚠️ Network or request error: {e}")
                    return None

                try:
                    data = response.json()

                    # Some APIs wrap data under "data"
                    if isinstance(data, dict) and "data" in data:
                        data = data["data"]

                    print("Returned data:", data)

                    if not data:
                        print(f"⚠️ No record found for UUID {new_uuid}")
                        return None

                    record = data[0] if isinstance(data, list) else data
                    print("Single record:", record)

                except ValueError as e:
                    print(f"⚠️ Invalid JSON response: {e}")
                    return None
                
                def get_val(key):
                    val = record.get(key)
                    if isinstance(val, (tuple, list)):
                        val = val[0] if val else None
                    if isinstance(val, str):
                        val = val.strip()
                        if val.lower() in ["", "null", "n/a", "none"]:
                            return None
                    return val

                instanceid = new_instance_id
                print(instanceid)

                # Save or update the record in local database
                b = Beneficiary(
                    record_id=get_val("_id"),
                    InForm_ID=get_val("_id"),
                    InstanceID=f"uuid:{get_val('_uuid')}",
                    IP_Name=get_val("IP_Name"),
                    Sector=get_val("Sector"),
                    Indicator=get_val("Indicator"),
                    Date=get_val("Date"),
                    Name=get_val("Name"),
                    ID_Number=get_val("ID_Number"),
                    Parent_ID=get_val("Parent_ID"),
                    Spouse_ID=get_val("Spouse_ID"),
                    Phone_Number=get_val("Phone_Number"),
                    Date_of_Birth=get_val("Date_of_Birth"),
                    Age=get_val("Age"),
                    Gender=get_val("Gender"),
                    Governorate=get_val("Governorate"),
                    Municipality=get_val("Municipality"),
                    Neighborhood=get_val("Neighborhood"),
                    Site_Name=get_val("Site_Name"),
                    Disability_Status=get_val("Disability_Status"),
                    Submission_Time=get_val("_submission_time"),
                    Deleted=get_val("Deleted"),
                    deleted_at=get_val("deleted_at"),
                    undeleted_at=get_val("undeleted_at"),
                    created_by=request.user,
                )
                b.save()
                messages.success(request, f"✅ New Beneficiary added & synced to InForm.", extra_tags="auto")

                
                
            else:
                messages.error(
                    request,
                    f"❌ Failed to add to InForm (status {response.status_code}): {response.text}",
                    extra_tags="auto"
                )
        # 🔹 Step 5: Update household relationships
            changes = assign_households()
            messages.success(request, f"✅ Beneficiary added & synced. Household updated for {changes} records.", extra_tags="auto")

        except Exception as e:
            messages.error(request, f"❌ Failed to connect to InForm: {str(e)}", extra_tags="auto")

        return redirect("/")


    ip_list = APIToken.objects.values_list('IP', flat=True).distinct().order_by('IP')
    sector_list = APIToken.objects.values_list('section', flat=True).distinct().order_by('section')

    context = {
        'ip_list': ip_list,
        'sector_list': sector_list,
        # include any other variables you already had in context
    }

    # GET request → render the Add form
    return render(request, "myproject/beneficiary_add.html", context)



# ---------- Manually run Assiogn_Households ---------------------------------------------------------------

@login_required
@role_required('Manager', 'Admin')
def assign_households_view(request):
    try:
        # Run your existing Python function
        result = assign_households()
        return JsonResponse({"status": "success", "message": "Households reassigned successfully."})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})
    

# ---------- SERIALIZERs ---------------------------------------------------------------


from rest_framework import viewsets, permissions
from .models import Beneficiary
from .serializers import BeneficiarySerializer
from rest_framework_simplejwt.authentication import JWTAuthentication
from .permissions import IsOwnerOrManager

class BeneficiaryViewSet(viewsets.ModelViewSet):
    serializer_class = BeneficiarySerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrManager]

    def get_queryset(self):
        user = self.request.user

        # Admin sees everything
        if user.is_superuser or user.groups.filter(name__in=["Admin", "Manager"]).exists():
            return Beneficiary.objects.all().order_by('-created_at')

        # Regular users see only their own records
        return Beneficiary.objects.filter(created_by=user).order_by('-created_at')

    def perform_create(self, serializer):
        # Automatically set created_by on record creation
        serializer.save(created_by=self.request.user)




#===============================BULK IMPORT ===================================
# 
# views.py (add after your other views)

# Global progress for bulk import (separate from sync_data)
bulk_progress_status = {
    "stage": "Not started",
    "progress": 0,
    "total": 0,
    "processed": 0,
    "created": 0,
    "failed": 0,
    "errors": [],
    "completed": False,
}

@login_required
def bulk_upload_page(request):
    """
    Render the page where user uploads Excel and picks IP.
    """
    ip_list = APIToken.objects.values_list('IP', flat=True).distinct().order_by('IP')
    context = {
        "ip_list": ip_list,
    }
    return render(request, "myproject/bulk_upload.html", context)


@login_required
def start_bulk_import(request):
    """
    Start background thread to parse uploaded Excel and import to InForm.
    Expects: POST with file field 'excel_file' and 'ip_name' (IP selected).
    """
    global bulk_progress_status

    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Only POST allowed."}, status=400)

    excel_file = request.FILES.get("excel_file")
    ip_name = request.POST.get("ip_name")

    if not excel_file:
        return JsonResponse({"status": "error", "message": "No file uploaded."}, status=400)
    if not ip_name:
        return JsonResponse({"status": "error", "message": "Please choose an Implementing Partner (IP)."}, status=400)

    # find apitoken for this ip
    apitoken_qs = APIToken.objects.filter(IP=ip_name)
    if not apitoken_qs.exists():
        return JsonResponse({"status": "error", "message": f"No APIToken found for IP {ip_name}"}, status=400)

    # choose first token for this IP (if multiple adjust as needed)
    api_token_obj = apitoken_qs.first()
    form_id = getattr(api_token_obj, "form_id", None)
    form_id_str = get_form_id_string(ip_name)
    token_value = getattr(api_token_obj, "token", None)

    if not form_id or not token_value:
        return JsonResponse({"status": "error", "message": "APIToken missing form_id or token for selected IP."}, status=400)

    # Save uploaded file temporarily to disk (openpyxl works with file path or file-like)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        for chunk in excel_file.chunks():
            tmp.write(chunk)
        tmp.flush()
        tmp.close()
    except Exception as e:
        try:
            os.unlink(tmp.name)
        except:
            pass
        return JsonResponse({"status": "error", "message": f"Failed saving uploaded file: {str(e)}"}, status=500)

    # reset progress
    bulk_progress_status = {
        "stage": "Queued",
        "progress": 0,
        "total": 0,
        "processed": 0,
        "created": 0,
        "failed": 0,
        "errors": [],
        "completed": False,
    }

    def run_import(path, form_id, token_value, ip_name):
        global bulk_progress_status
        try:
            bulk_progress_status.update({"stage": "Reading Excel", "progress": 2})
            # Read workbook and first sheet
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            sheet = wb.active

            # Read header row
            rows = sheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                bulk_progress_status.update({
                    "stage": "❌ Excel contains no rows",
                    "progress": 100,
                    "completed": True,
                    "errors": ["Excel contains no rows"]
                })
                return

            header_map = []
            for idx, col in enumerate(header):
                if col is None:
                    header_map.append(f"col_{idx}")
                else:
                    header_map.append(str(col).strip())

            # expected InForm field names - you can add more mapping here if your Excel uses different names
            expected_fields = [
                "IP_Name","Sector","Indicator","Date","Name","ID_Number","Parent_ID","Spouse_ID",
                "Phone_Number","Date_of_Birth","Age","Gender","Governorate","Municipality","Neighborhood",
                "Site_Name","Disability_Status"
            ]

            # Read all rows into list of dicts
            records = []
            for r in rows:
                if not any([c is not None and str(c).strip() != "" for c in r]):
                    continue  # skip empty rows
                rec = {}
                for idx, value in enumerate(r):
                    col_name = header_map[idx]
                    # Basic normalization: if header matches expected field case-insensitively, map it
                    mapped = None
                    for f in expected_fields:
                        if col_name.lower() == f.lower():
                            mapped = f
                            break
                    # Additional common header aliases
                    if not mapped:
                        alias_map = {
                            "full name": "Name", "name": "Name",
                            "id": "ID_Number", "id number": "ID_Number", "national id": "ID_Number",
                            "phone": "Phone_Number", "phone number": "Phone_Number",
                            "dob": "Date_of_Birth", "date of birth": "Date_of_Birth",
                            "age": "Age", "gender": "Gender",
                            "governorate": "Governorate", "municipality": "Municipality",
                            "neighborhood": "Neighborhood", "site": "Site_Name",
                            "disability": "Disability_Status", "sector": "Sector", "ip_name": "IP_Name", "ip": "IP_Name",
                            "indicator": "Indicator", "date": "Date"
                        }
                        if col_name.strip().lower() in alias_map:
                            mapped = alias_map[col_name.strip().lower()]
                    # fallback: use header name directly (if safe)
                    if not mapped:
                        mapped = col_name
                    rec[mapped] = value
                records.append(rec)

            total = len(records)
            bulk_progress_status.update({"stage": "Preparing to post", "progress": 5, "total": total})
            if total == 0:
                bulk_progress_status.update({
                    "stage": "❌ No valid rows found in Excel",
                    "progress": 100,
                    "completed": True,
                    "errors": ["No valid rows found"]
                })
                return

            # POST each record to InForm API
            created = 0
            failed = 0
            processed = 0
            errors = []

            for idx, rec in enumerate(records, start=1):
                processed += 1
                try:
                    # ensure IP_Name in rec
                    if not rec.get("IP_Name"):
                        rec["IP_Name"] = ip_name
                    # Check duplicates locally before posting -----------------
                    # 🚫 Duplicate ID_Number check (before sending to InForm)
                    existing = Beneficiary.all_objects.filter(ID_Number=str(rec.get("ID_Number")).strip()).exists()
                    if existing:
                        failed += 1
                        errors.append(f"Row {idx}: ID_Number already exists ({rec.get('ID_Number')}). Skipped.")
                        bulk_progress_status.update({
                            "stage": f"Duplicate detected in row {idx}",
                            "progress": int((processed / max(total,1)) * 95),
                            "processed": processed,
                            "failed": failed,
                            "created": created,
                            "errors": errors[:10],
                            "completed": False
                        })
                        continue  # skip sending to InForm

                    # create instanceID for the submission (uuid)
                    import uuid
                    new_uuid = uuid.uuid4()
                    instance_id = f"uuid:{new_uuid}"

                    payload = {
                        "id": form_id_str,
                        "submission": {
                            **{k: ("" if v is None else v) for k,v in rec.items()},
                            "created_by": request.user,
                            "meta": {"instanceID": instance_id}
                        }
                    }
                    headers = {
                        "Authorization": f"Token {token_value}",
                        "Content-Type": "application/json"
                    }

                    # Post to InForm submission endpoint (adjust URL pattern to your environment)
                    inform_api_url = f"https://data.inform.unicef.org/unicefstateofpalestine/submission"
                    resp = requests.post(inform_api_url, headers=headers, json=payload, timeout=30)
                    if resp.status_code not in [200,201]:
                        # record failure
                        failed += 1
                        errors.append(f"Row {idx}: post failed status {resp.status_code} - {resp.text[:200]}")
                    else:
                        created += 1

                        # Immediately fetch the created record by UUID (safer than relying on post response)
                        uuid_str = str(new_uuid) if not isinstance(new_uuid, str) else new_uuid
                        query = json.dumps({"_uuid": uuid_str})
                        print(uuid_str)
                        fetch_url = f"https://data.inform.unicef.org/api/v1/data/{form_id}.json?query={query}"
                        try:
                            r2 = requests.get(fetch_url, headers={"Authorization": f"Token {token_value}"}, timeout=30)
                            if r2.status_code == 200:
                                data = r2.json()
                                if isinstance(data, dict) and "data" in data:
                                    data = data["data"]
                                record = data[0] if isinstance(data, list) and data else (data if isinstance(data, dict) else None)
                                if record:
                                    print(record)
                                    # Save/update local Beneficiary (use same logic as in beneficiary_add)
                                    def get_val(key):
                                        v = record.get(key)
                                        if isinstance(v, (tuple, list)):
                                            v = v[0] if v else None
                                        if isinstance(v, str):
                                            v = v.strip()
                                            if v.lower() in ["", "null", "n/a", "none"]:
                                                return None
                                        return v

                                    Beneficiary.all_objects.update_or_create(
                                        record_id=get_val("_id"),
                                        defaults={
                                            "record_id": get_val("_id"),
                                            "InForm_ID": get_val("_id"),
                                            "InstanceID": f"uuid:{get_val('_uuid')}",
                                            "IP_Name": get_val("IP_Name"),
                                            "Sector": get_val("Sector"),
                                            "Indicator": get_val("Indicator"),
                                            "Date": get_val("Date"),
                                            "Name": get_val("Name"),
                                            "ID_Number": get_val("ID_Number"),
                                            "Parent_ID": get_val("Parent_ID"),
                                            "Spouse_ID": get_val("Spouse_ID"),
                                            "Phone_Number": get_val("Phone_Number"),
                                            "Date_of_Birth": get_val("Date_of_Birth"),
                                            "Age": get_val("Age"),
                                            "Gender": get_val("Gender"),
                                            "Governorate": get_val("Governorate"),
                                            "Municipality": get_val("Municipality"),
                                            "Neighborhood": get_val("Neighborhood"),
                                            "Site_Name": get_val("Site_Name"),
                                            "Disability_Status": get_val("Disability_Status"),
                                            "Submission_Time": get_val("_submission_time"),
                                            "Deleted": get_val("Deleted"),
                                            "deleted_at": get_val("deleted_at"),
                                            "undeleted_at": get_val("undeleted_at"),
                                            "created_by": request.user,
                                        }
                                    )
                        except Exception as e_fetch:
                            # non-fatal: log but continue
                            errors.append(f"Row {idx}: fetch after post failed: {str(e_fetch)[:200]}")

                    # update progress per-record
                    percent = int((processed / max(total,1)) * 95)  # keep final 95 for finishing tasks
                    bulk_progress_status.update({
                        "stage": f"Processing row {processed}/{total}",
                        "progress": percent,
                        "total": total,
                        "processed": processed,
                        "created": created,
                        "failed": failed,
                        "errors": errors[:10],
                        "completed": False
                    })

                except Exception as e_row:
                    failed += 1
                    errors.append(f"Row {idx}: unexpected error: {str(e_row)[:200]}")
                    bulk_progress_status.update({
                        "stage": f"Error processing row {idx}",
                        "progress": int((processed / max(total,1)) * 95),
                        "errors": errors[:10],
                        "completed": False
                    })
                    continue

            # after all rows posted, run household assignment
            bulk_progress_status.update({"stage": "Running household generation...", "progress": 96})
            try:
                changes = assign_households()
                # add a friendly message
                bulk_progress_status.setdefault("errors", [])
                bulk_progress_status["errors"].append(f"Household generation changed {changes} records.")
            except Exception as e_h:
                bulk_progress_status.setdefault("errors", []).append(f"Household generation failed: {str(e_h)[:200]}")

            # finalize
            bulk_progress_status.update({
                "stage": "✅ Import completed",
                "progress": 100,
                "total": total,
                "processed": processed,
                "created": created,
                "failed": failed,
                "errors": errors,
                "completed": True
            })

        except Exception as e:
            bulk_progress_status.update({
                "stage": f"❌ Unexpected error: {str(e)[:200]}",
                "progress": 100,
                "completed": True,
                "errors": [str(e)]
            })
        finally:
            # cleanup
            try:
                os.unlink(path)
            except:
                pass

    # launch thread
    t = threading.Thread(target=run_import, args=(tmp.name, form_id, token_value, ip_name), daemon=True)
    t.start()

    return JsonResponse({"status": "started"})


@login_required
def get_bulk_progress(request):
    """
    Return current bulk import progress as JSON.
    """
    global bulk_progress_status
    return JsonResponse(bulk_progress_status)


def autocomplete(request):
    field = request.GET.get("field")
    query = request.GET.get("q", "")

    if not field:
        return JsonResponse([], safe=False)

    # Get unique values for that field
    values = (
        Beneficiary.objects
        .filter(**{f"{field}__icontains": query})
        .values_list(field, flat=True)
        .distinct()
    )

    return JsonResponse(list(values[:15]), safe=False)
    


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_user_info(request):
    user = request.user
    return Response({
        'username': user.username,
        'firstname': user.first_name,
        'lastname': user.last_name,
        'email': user.email,
        'groups': list(user.groups.values_list('name', flat=True))
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_duplicate(request, id_number):
    exists = Beneficiary.objects.filter(id_number=id_number).exists()
    return Response({"exists": exists})


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import Beneficiary


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def distinct_ip_names(request):
    ip_names = (
        Beneficiary.objects
        .exclude(ipname__isnull=True)
        .exclude(ipname__exact='')
        .values_list('ipname', flat=True)
        .distinct()
        .order_by('ipname')
    )
    return Response({"ip_names": list(ip_names)})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def distinct_sectors(request):
    sectors = (
        Beneficiary.objects
        .exclude(sector__isnull=True)
        .exclude(sector__exact='')
        .values_list('sector', flat=True)
        .distinct()
        .order_by('sector')
    )
    return Response({"sectors": list(sectors)})


