import time
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.conf import settings
import requests
from .models import Beneficiary, APIToken, Supply
from datetime import datetime, timedelta
from django.core.paginator import Paginator
from django.http import JsonResponse,HttpResponse, HttpResponseForbidden, Http404
from django.utils.timezone import localtime
from .forms import APITokenForm
from datetime import datetime, date
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy, reverse
from .forms import StyledPasswordChangeForm
import json
import logging
from django.db.models import Q
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.db.models.functions import Lower
from audit.models import AuditLog 
from audit.views import log_user_action
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt  # only if using non-ajax or external POSTs — prefer csrf
from django.db.models.functions import TruncMonth
# wherever you create beneficiary from get_val(...)
from .households import assign_households
from django.db.models import Case, When, Value, CharField, F
from django.db.models.functions import Lower
from openpyxl import Workbook, load_workbook  # used to read Excel files
from openpyxl.styles import Font
import tempfile
from django.utils.dateparse import parse_datetime
from django.utils import timezone
from rest_framework.response import Response

## ----------------Shared Variables --------------------------------------

API_URL = settings.API_URL
INFORM_API_URL = settings.INFORM_API_URL
API_TOKEN = settings.API_TOKEN

## ----------------Permiossion Denied, Role required ----------------------

from django.core.exceptions import PermissionDenied

def role_required(*roles):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)
            if request.user.groups.filter(name__in=roles).exists():
                return view_func(request, *args, **kwargs)
            raise PermissionDenied
        return wrapper
    return decorator

def log_user_action(request, action, model_name=None, record_id=None, changes=None):
    user = request.user if request.user.is_authenticated else None
    AuditLog.objects.create(
        user=user,
        action=action,
        model_name=model_name,
        record_id=record_id,
        changes=changes or {},
        timestamp=timezone.now()
    )



@login_required
def home(request):
    # Gender distribution
    # Aggregate gender counts

    # Get distinct sectors for the filter dropdown
    sector_list = (
        Beneficiary.objects
        .exclude(Sector__isnull=True)
        .exclude(Sector='')
        .values_list('Sector', flat=True)
        .distinct()
        .order_by('Sector')
    )

    selected_sector = request.GET.get('sector', '')

    base_qs = Beneficiary.objects.all()

    if selected_sector:
        base_qs = base_qs.filter(Sector=selected_sector)


    gender_data = base_qs.values('Gender').annotate(total=Count('id'))

    # Initialize counters
    gender_counts = {'Male': 0, 'Female': 0, 'Unknown': 0}

    # Normalize and sum
    for g in gender_data:
        gender = (g['Gender'] or '').strip().lower()
        if gender in ['m', 'male']:
            gender_counts['Male'] += g['total']
        elif gender in ['f', 'female']:
            gender_counts['Female'] += g['total']
        else:
            gender_counts['Unknown'] += g['total']

    # Prepare for JS
    gender_labels = ['Male', 'Female', 'Unknown']
    gender_values = [gender_counts['Male'], gender_counts['Female'], gender_counts['Unknown']]

    
    # Age groups (example buckets)
    age_buckets = {
        '0-17': base_qs.filter(Age__lte=17).count(),
        '18-35': base_qs.filter(Age__gte=18, Age__lte=35).count(),
        '36-60': base_qs.filter(Age__gte=36, Age__lte=60).count(),
        '60+': base_qs.filter(Age__gte=61).count(),
    }

    # Sector
    sector_data = base_qs.values('IP_Name').annotate(total=Count('id'))
    sector_labels = [s['IP_Name'] or 'Unknown' for s in sector_data]
    sector_values = [s['total'] for s in sector_data]

    
    # Aggregate submissions by month and IP_Name
    submissions = (
        base_qs
        .annotate(month=TruncMonth('Submission_Time'))
        .values('month', 'IP_Name')
        .annotate(total=Count('id'))
        .order_by('month', 'IP_Name')
    )

    # All IP names in the dataset
    ip_names = sorted({s['IP_Name'] or "Unknown" for s in submissions})

    # Sorted months
    submission_months = sorted({s['month'] for s in submissions if s['month']})
    submission_labels = [m.strftime("%b %Y") for m in submission_months]

    # Map month objects to their index
    month_index_map = {m: i for i, m in enumerate(submission_months)}

    # Initialize datasets
    datasets = {ip: [0] * len(submission_labels) for ip in ip_names}

    # Fill datasets
    for s in submissions:
        if s['month'] is None:
            continue
        ip = s['IP_Name'] or "Unknown"
        idx = month_index_map[s['month']]
        datasets[ip][idx] = s['total']

    # Chart.js datasets
    submission_datasets = [{"label": ip, "data": datasets[ip]} for ip in ip_names]


    # Governorates
    
    gov_data = (
        base_qs
        .annotate(
            gov_raw=Case(
                When(Governorate__isnull=True, then=Value('Unknown')),
                When(Governorate='', then=Value('Unknown')),
                default=F('Governorate'),
                output_field=CharField(),
            ),
            gov_cleaned=Lower(
                Case(
                    When(Governorate__isnull=True, then=Value('Unknown')),
                    When(Governorate='', then=Value('Unknown')),
                    default=F('Governorate'),
                    output_field=CharField(),
                )
            )
        )
        .values('gov_cleaned')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )


    gov_labels = [
        g['gov_cleaned'].capitalize() if g['gov_cleaned'] != "unknown" else "Unknown"
        for g in gov_data
    ]

    gov_values = [g['total'] for g in gov_data]



    last_sync_record = Beneficiary.all_objects.order_by('-created_at').first()

    

    context = {
        'sector_list': sector_list,
        'selected_sector': selected_sector,
        "gender_labels": gender_labels,
        "gender_data": gender_values,
        "age_labels": list(age_buckets.keys()),
        "age_data": list(age_buckets.values()),
        "sector_labels": sector_labels,
        "sector_data": sector_values,
        "submission_labels": submission_labels,
        "submission_datasets": submission_datasets,
        "gov_labels": gov_labels,
        "gov_data": gov_values,
        
        "last_sync_record": Beneficiary.all_objects.order_by('-created_at').first(),
        "last_sync_time" : localtime(last_sync_record.created_at).strftime("%Y-%m-%d %H:%M") if last_sync_record else "1900-01-01 00:00",
    
    }
    return render(request, "myproject/home.html", context)

     
    
    return render(request, 'myproject/home.html', {'last_sync_time': last_sync_time})


def get_item(dictionary, key):
    if isinstance(dictionary, dict):
        return dictionary.get(key, "")
    return ""


import os, threading, time

# global variable for progress
progress_status = {"stage": "Fetching Data", "progress": 0}

@login_required
@role_required('Manager', 'Admin')
def sync_data(request):
    global progress_status

    
    
    if request.method != 'POST':
        return JsonResponse({"status": "error", "message": "Invalid request method. Only POST allowed.", "extra_tags":"auto"})

    # Reset before new run
    progress_status = {"stage": "Starting sync ...", "progress": 1, "completed": False}

    def run_sync():
        global progress_status
        tokens = APIToken.objects.all()
        if not tokens.exists():
            progress_status = {"stage": "❌ No API tokens available.", "progress": 0, "completed": True}
            return

        total_fetched = 0
        new_count = 0
        duplicate_count = 0
        invalid_date_count = 0

        # iterate tokens but do NOT let one token failure abort everything
        for token in tokens:
            token_id = getattr(token, "form_id", str(token.pk))
            page = 1
            page_size = 10000

            last_sync_record = Beneficiary.all_objects.order_by('-created_at').first()
            last_sync_time = (
                localtime(last_sync_record.created_at).strftime("%Y-%m-%d %H:%M")
                if last_sync_record
                else "1900-01-01 00:00"
            )

            # token-level try so we continue to next token on failure
            try:
                while True:
                    url = f"{API_URL}{token.form_id}.json?page={page}&page_size={page_size}"
                    headers = {"Authorization": f"Token {API_TOKEN}", "Content-Type": "application/json"}

                    try:
                        response = requests.get(url, headers=headers, timeout=30)
                    except requests.RequestException as e:
                        # network/timeout error — report and continue to next token
                        progress_status = {
                            "stage": f"⚠️ Network error for token {token_id}: {str(e)}",
                            "progress": 0,
                            "completed": False,
                        }
                        break  # stop pages for this token, move to next token

                    if response.status_code == 401 or response.status_code == 403:
                        # unauthorized for this token — skip token
                        progress_status = {
                            "stage": f"⚠️ Auth failed for token {token_id} (HTTP {response.status_code}) — skipping token",
                            "progress": 0,
                            "completed": False,
                        }
                        break

                    if response.status_code != 200:
                        # other non-200 — log and stop paging this token
                        progress_status = {
                            "stage": f"⚠️ Failed fetching token {token_id} page {page} (HTTP {response.status_code})",
                            "progress": 0,
                            "completed": False,
                        }
                        break

                    # parse JSON safely
                    try:
                        data = response.json()
                        # ✅ Stop when there’s no data or an empty page
                        if not data or (isinstance(data, dict) and not data.get("data") and not data.get("results")):
                            break
                        # If API wraps data under "data"
                        if isinstance(data, dict) and "data" in data:
                            data = data["data"]

                        if len(data) == 0:
                            break

                        total_fetched += len(data)
                    except ValueError:
                        progress_status = {
                            "stage": f"⚠️ Invalid JSON for token {token_id} page {page} — skipping token",
                            "progress": 0,
                            "completed": False,
                        }
                        break

                    if not data:
                        # no more pages
                        break

                    # increment totals for reporting
                    #total_fetched += len(data)

                    for idx, item in enumerate(data, start=1):
                        def get_val(key):
                            val = item.get(key)
                            if isinstance(val, (tuple, list)):
                                val = val[0] if val else None
                            if isinstance(val, str):
                                val = val.strip()
                                if val.lower() in ["", "null", "n/a", "none"]:
                                    return None
                            return val

                        # progress update (per-record)
                        progress_status = {
                            "stage": f"IP: {get_val('IP_Name')} | Form: {token_id} | Page: {page} | Record:({idx}/{len(data)})",
                            # show progress up to 98 while running; final will be 100 when done
                            "progress": min(98, int((idx / max(len(data), 1)) * 100)),
                            "new_count": new_count,
                            "duplicate_count": duplicate_count,
                            "total_fetched": total_fetched,
                            "completed": False,
                        }
                        InForm_ID = get_val("_id")
                        

                        #last_synced = get_val('_submission_time')
                        # Try to find an existing beneficiary with this InForm_ID
                        
                        #new_rec = last_synced > last_sync_time
                        #print(f"New record check: {new_rec} for submission time {last_synced} vs last sync {last_sync_time}")
                        #if new_rec:
                        if Beneficiary.all_objects.filter(record_id=InForm_ID).exists():
                            continue  # skip duplicates
                            # Create new record only if not found or InForm_ID is empty
                        Beneficiary.all_objects.update_or_create(
                            record_id=get_val("_id"),
                            defaults={
                                "record_id": get_val("_id"),
                                "InForm_ID": get_val("_id"),
                                "InstanceID": f"uuid:{get_val('_uuid')}",
                                "IP_Name": get_val("IP_Name"),
                                "Sector": get_val("Sector"),
                                "Indicator": get_val("Indicator"),
                                "Date": get_val("Date"),
                                "Name": get_val("Name"),
                                "ID_Number": get_val("ID_Number"),
                                "Parent_ID": get_val("Parent_ID"),
                                "Spouse_ID": get_val("Spouse_ID"),
                                "Phone_Number": get_val("Phone_Number"),
                                "Date_of_Birth": get_val("Date_of_Birth"),
                                "Age": get_val("Age"),
                                "Gender": get_val("Gender"),
                                "Governorate": get_val("Governorate"),
                                "Municipality": get_val("Municipality"),
                                "Neighborhood": get_val("Neighborhood"),
                                "Site_Name": get_val("Site_Name"),
                                "Disability_Status": get_val("Disability_Status"),
                                "Submission_Time": get_val("_submission_time"),
                                "Deleted": get_val("Deleted"),
                                "deleted_at": get_val("deleted_at"),
                                "undeleted_at": get_val("undeleted_at"),
                                
                            }
                        )
                        new_count += 1

                    # next page for this token
                    # ✅ Move to next page only if this page was full
                    if len(data) < page_size:
                        # means this was the last page
                        break

                    page += 1

            except Exception as e_token:
                # catch unexpected exceptions for this token and continue to the next
                progress_status = {
                    "stage": f"❌ Unexpected error while processing token {token_id}: {str(e_token)}",
                    "progress": 0,
                    "completed": False,
                }
                # continue to next token
                continue

        # After all tokens processed, assign households and produce final summary
        try:
            changes = assign_households()
            messages.success(request, f"Household assignment updated for {changes} records.", extra_tags="auto")
        except Exception as e_house:
            changes = 0
            # log or include house error in the status
            progress_status = {
                "stage": f"⚠️ Household assignment error: {str(e_house)}",
                "progress": 100,
                "completed": True,
            }
        messages.success(request, f"Household assignment updated for {changes} records.", extra_tags="auto")

        progress_status = {
            "stage": "✅ Sync completed",
            "progress": 100,
            "new_count": new_count,
            "duplicate_count": duplicate_count,
            "invalid_date_count": invalid_date_count,
            "total_fetched": total_fetched,
            "household_changes": changes,
            "completed": True,
            "extra_tags":"auto"
        }

    thread = threading.Thread(target=run_sync, daemon=True)
    thread.start()

    return JsonResponse({"status": "started"})
    
    

    


def get_progress(request):
    global progress_status
    return JsonResponse(progress_status)


@login_required
def inform_data_view(request):

    # Admins/Managers see everything
    if request.user.is_superuser or request.user.groups.filter(name__in=["Manager", "Admin"]).exists():
        submissions = Beneficiary.objects.all()
        #print(f"User {request.user.username} sees all records.")
    # Normal staff only see their own records
    else:
        submissions = Beneficiary.objects.filter(created_by=request.user.username)
        #print(f"User {request.user.username} sees only their records.")

    # --- Sorting ---
    sort_field = request.GET.get('sort', '-created_at')  # default sort
    # Ensure the sort field is valid
    valid_fields = [f.name for f in Beneficiary._meta.fields]
    if sort_field.lstrip('-') in valid_fields:
        submissions = submissions.order_by(sort_field)
    else:
        submissions = submissions.order_by('-created_at')

    # --- Filtering ---
    filter_fields = ['IP_Name','Sector','Indicator','Date','Name','ID_Number','Parent_ID','Spouse_ID','Phone_Number',
                     'Date_of_Birth','Age','Gender','Governorate','Household_ID']
    
    from django.db.models import Q

    operator = request.GET.get('operator', 'AND').upper()

    # Build Q objects for each filter field
    q_objects = []
    for field in filter_fields:  # filter_fields should be list of (field_name, label)
        value = request.GET.get(field[0], '').strip()
        if value:
            q_objects.append(Q(**{f"{field[0]}__icontains": value}))

    
    # Apply AND / OR
    qs = Beneficiary.objects.all()
    if operator == 'AND':
        for q in q_objects:
            qs = qs.filter(q)
    #elif operator == 'OR' and q_objects:
    #    combined_q = Q()
    #    for q in q_objects:
    #        combined_q |= q
    #    qs = qs.filter(combined_q)

    for field in filter_fields:
        value = request.GET.get(field)
        if value:
            submissions = submissions.filter(**{f"{field}__icontains": value})
    

    cnt = Beneficiary.objects.count()  # total in DB

    # --- Pagination ---
    paginator = Paginator(submissions, 18)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Fields for filter inputs (friendly labels)
    fields = [
        ('IP_Name','IP Name'),
        ('Sector','Sector'),
        ('Indicator','Indicator'),
        ('Date','Date'),
        ('Name','Name'),
        ('ID_Number','ID Number'),
        ('Parent_ID','Parent_ID'),
        ('Spouse_ID','Spouse_ID'),
        ('Phone_Number','Phone Number'),
        ('Date_of_Birth','Date of Birth'),
        ('Age','Age'),
        ('Gender','Gender'),
        ('Governorate','Governorate'),
        ('Household_ID','Household_ID'),
    ]

    context = {
        'submissions': page_obj,
        'filter_fields': fields,
        'request': request,  # for pagination links
        'total_records': submissions.count(),  # total after filtering
        'total_all_records': cnt,  # total in DB
        'current_sort': sort_field,  # current sorting
    }

    # AJAX request returns only table
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'myproject/inform_table.html', context)

    return render(request, 'myproject/inform_data.html', context)




@login_required
@role_required('Manager', 'Admin')
def manage_tokens(request):
    if request.method == 'POST':
        form = APITokenForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Token saved successfully.", extra_tags="auto")
            return redirect('manage_tokens')
        else:
            messages.error(request, "❌ Please correct the errors below.", extra_tags="auto")
    else:
        form = APITokenForm()

    tokens = APIToken.objects.all().order_by('-id')  # Optional: newest on top
    return render(request, 'myproject/manage_tokens.html', {
        'form': form,
        'tokens': tokens
    })

@login_required
@role_required('Manager', 'Admin')
def delete_token(request, pk):
    token = get_object_or_404(APIToken, pk=pk)
    token.delete()
    messages.success(request, "API token deleted successfully.", extra_tags="auto")
    return redirect('manage_tokens')  # Redirect back to the manage tokens page







def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, "You have successfully logged in.", extra_tags="auto")
            # NOW user is a real User, safe to log
            AuditLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action="LOGIN",
                model_name="User",
                record_id=str(request.user.id) if request.user.is_authenticated else None,
                changes={},  # or any relevant info
                description="User Logged In",
                timestamp=timezone.now()
            )

            return redirect('home')  # Redirect to your desired page
        else:
            messages.error(request, "Invalid username or password.", extra_tags="auto")
    else:
        form = AuthenticationForm()

    
    context = {
        'form': form,
    }
    return render(request, 'myproject/login.html', context)

def user_logout(request):
    
    AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action="LOGOUT",
            model_name="User",
            record_id=str(request.user.id) if request.user.is_authenticated else None,
            changes={},  # or any relevant info
            description="User Logged Out",
            timestamp=timezone.now()
        )

    logout(request)
    
    messages.success(request, "You have been logged out.", extra_tags="auto")
    return redirect("login")  # Replace with your login URL name



def user_register(request):
    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if password == confirm_password:
            if User.objects.filter(username=username).exists():
                messages.error(request, "Username already exists.", extra_tags="auto")
            elif User.objects.filter(email=email).exists():
                messages.error(request, "Email is already registered.", extra_tags="auto")
            else:
                user = User.objects.create_user(username=username, email=email, first_name=first_name, last_name=last_name, password=password)
                user.is_active = True  # 🔑 make sure the account is active
                user.save()
                #print(user.password)
                messages.success(request, "Account created successfully! Please log in.", extra_tags="auto")
                return redirect('login')
        else:
            messages.error(request, "Passwords do not match.", extra_tags="auto")
    return render(request, 'myproject/register.html')


@login_required
def profile(request):
    if request.method == "POST":
        user = request.user
        user.first_name = request.POST.get("first_name")
        user.last_name = request.POST.get("last_name")
        user.email = request.POST.get("email")
        user.save()
        messages.success(request, "User profile updated successfully!", extra_tags="auto")
        return redirect("home")
    return render(request, "myproject/profile.html")

@login_required
def password_change(request):
    if request.method == "POST":
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Keep user logged in after password change
            update_session_auth_hash(request, user)
            messages.success(request, "✅ Your password was successfully updated!", extra_tags="auto")
            return redirect("home")  # redirect to profile or dashboard
        else:
            messages.error(request, "❌ Please correct the errors below.", extra_tags="auto")
    else:
        form = PasswordChangeForm(request.user)

    return render(request, "myproject/password_change.html", {"form": form})


class CustomPasswordChangeView(PasswordChangeView):
    form_class = StyledPasswordChangeForm
    template_name = "myproject/change_password.html"
    success_url = reverse_lazy('home')


# ------- ID_Number Duplication Section -----------------------------------------------------------

logger = logging.getLogger(__name__)
@login_required
def find_duplicates(request):
    """
    Page that finds duplicates in Beneficiary table grouped by id_number,
    and displays each duplicate set in its own table.
    """
    # Adjust `id_number` field name if your model uses different casing.
    # Group by id_number and keep groups with count>1

    # Admins/Managers see everything
    if request.user.is_superuser or request.user.groups.filter(name__in=["Manager", "Admin"]).exists():
        submissions = Beneficiary.objects.all()

    # Normal staff only see their own records
    else:
        submissions = Beneficiary.objects.filter(created_by=request.user)


    total_records = 0
    duplicate_groups = (
        submissions
        .values("ID_Number")
        .annotate(count=Count("ID_Number"))
        .filter(count__gt=1)
        .order_by("-count")
    )

    # Build a list of groups where each group contains the actual records
    groups = []
    for g in duplicate_groups:
        ID_Number = g["ID_Number"]
        records = list(Beneficiary.objects.filter(ID_Number=ID_Number).order_by('-created_at'))
        groups.append({
            "ID_Number": ID_Number,
            "count": g["count"],
            "records": records,
        })
        total_records += len(records)

    context = {
        "groups": groups,
        "total_duplicates": len(groups),
        "total_records": total_records,
    }
    return render(request, "myproject/duplicates.html", context)


@require_POST
@login_required
def delete_inform(request, pk):
   
    if not request.user.is_authenticated:
        return HttpResponseForbidden("Authentication required.")

    
    # Get target record
    obj = get_object_or_404(Beneficiary, pk=pk)


    try:
        # Delete the local DB record
        obj.soft_delete()
        messages.success(request, "✅ Record has been deleted successfully and synced to InForm.", extra_tags="auto")
    except Exception as e:
            messages.error(request, f"❌ Record could not be deleted: {str(e)}", extra_tags="auto")

    return JsonResponse({"success": True, "deleted_pk": pk})

@require_POST
@login_required
def delete_duplicate(request, pk):
    
    if not request.user.is_authenticated:
        return HttpResponseForbidden("Authentication required.")

    # Get target record
    obj = get_object_or_404(Beneficiary, pk=pk)
    

    try:
        # Delete the local DB record
        obj.soft_delete()
        messages.success(request, "✅ Record has been deleted successfully and synced to InForm.", extra_tags="auto")
    except Exception as e:
            messages.error(request, f"❌ Record could not be deleted: {str(e)}", extra_tags="auto")

    return JsonResponse({"success": True, "deleted_pk": pk})


# ---------- Beneficiary Details page ---------------------------------------------------------------

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
import requests, uuid
from .models import Beneficiary

def safe_value(val):
    """Convert empty strings to None for API submission."""
    return val if val not in ["", None] else None

def get_form_id(ip_name):
    """Return correct Form_ID_String based on IP_Name."""
    mapping = {
        "AEI": "8794",
        "UNRWA": "9190",
        "WeWorld": "9288",
        "MedGlobal": "8793",
        "RRM": "8927",
        "NECC": "9094",
        "Taawon": "10000",
        "Maan": "10319",
    }
    return mapping.get(ip_name, "00000")

def get_form_id_string(ip_name):
    """Return correct Form_ID_String based on IP_Name."""
    mapping = {
        "AEI": "Beneficiary_Database_Form_Template",
        "UNRWA": "Beneficiary_Database_Form_Template_-_UNRWA",
        "WeWorld": "Beneficiary_Database_Form_Template_-_We_World",
        "MedGlobal": "azPSg7utHm6x4kNyfpdgCj",
        "RRM": "Beneficiary_Database_Form_Template_-_RRM",
        "NECC": "Beneficiary_Database_Form_Template_-_NECC",
        "Taawon": "Beneficiary_Database_Form_Template_-_Taawon",
        "Maan": "Beneficiary_Database_Form_Template_-_Maan",
    }
    return mapping.get(ip_name, "Beneficiary_Database_Form_Template_-_Default")

@login_required
def beneficiary_details(request, pk):

    record = get_object_or_404(Beneficiary, pk=pk)
    user = request.user if request.user.is_authenticated else None

    if request.method == "POST":
        try:
            # -----------------------------
            # UPDATE LOCAL FIELDS
            # -----------------------------
            record.IP_Name = safe_value(request.POST.get("IP_Name"))
            record.Sector = safe_value(request.POST.get("Sector"))
            record.Indicator = safe_value(request.POST.get("Indicator"))
            record.Date = safe_value(request.POST.get("Date"))
            record.Name = safe_value(request.POST.get("Name"))
            record.ID_Number = safe_value(request.POST.get("ID_Number"))
            record.Parent_ID = safe_value(request.POST.get("Parent_ID"))
            record.Spouse_ID = safe_value(request.POST.get("Spouse_ID"))
            record.Phone_Number = safe_value(request.POST.get("Phone_Number"))
            record.Date_of_Birth = safe_value(request.POST.get("Date_of_Birth"))
            record.Age = safe_value(request.POST.get("Age"))
            record.Gender = safe_value(request.POST.get("Gender"))
            record.Governorate = safe_value(request.POST.get("Governorate"))
            record.Municipality = safe_value(request.POST.get("Municipality"))
            record.Neighborhood = safe_value(request.POST.get("Neighborhood"))
            record.Site_Name = safe_value(request.POST.get("Site_Name"))
            record.Disability_Status = safe_value(request.POST.get("Disability_Status"))
            record.HH_Members = safe_value(request.POST.get("HH_Members"))
            record.Marital_Status = safe_value(request.POST.get("Marital_Status"))
            record.Supply_Type = safe_value(request.POST.get("Supply_Type"))
            record.Benefit_Date = safe_value(request.POST.get("Benefit_Date"))
            record.updated_at = timezone.now()
            record.updated_by = user

            # -----------------------------
            # SAVE FIRST (without eligibility)
            # -----------------------------
            record.save()

            # -----------------------------
            # CALCULATE ELIGIBILITY (CORRECT)
            # -----------------------------
            record.Eligiblity = calculate_eligibility(record)
            record.save(update_fields=["Eligiblity"])

            # -----------------------------
            # UPDATE HOUSEHOLDS
            # -----------------------------
            changes = assign_households()

            messages.success(
                request,
                f"✅ Beneficiary updated successfully. Household updated for {changes} records.",
                extra_tags="auto",
            )

        except Exception as e:
            messages.error(
                request,
                f"❌ Failed to update beneficiary: {str(e)}",
                extra_tags="auto",
            )

        return redirect("beneficiary_details", pk=record.pk)

    # -----------------------------
    # GET REQUEST (FORM DATA)
    # -----------------------------
    ip_list = (
        APIToken.objects
        .values_list("IP", flat=True)
        .distinct()
        .order_by("IP")
    )

    sector_list = (
        APIToken.objects
        .values_list("section", flat=True)
        .distinct()
        .order_by("section")
    )

    supply_types = (
        Supply.objects
        .exclude(supply_type__isnull=True)
        .exclude(supply_type="")
        .values_list("supply_type", flat=True)
        .distinct()
        .order_by("supply_type")
    )

    context = {
        "record": record,
        "ip_list": ip_list,
        "sector_list": sector_list,
        "supply_types": supply_types,
    }

    return render(request, "myproject/beneficiary_details.html", context)


# ---------- Household Details page ---------------------------------------------------------------


@login_required
def household_details(request, household_id):
    # Get all members in this household
    members = Beneficiary.objects.filter(Household_ID=household_id)

    if not members.exists():
        raise Http404("Household not found")

    # Identify potential head (Parent_ID is empty, usually the head)
    head = next((m for m in members if not m.Parent_ID), members.first())

    # Build relationship mapping
    relationships = []
    for member in members:
        relation = "Head" if member == head else "Spouse" if member.Spouse_ID == head.ID_Number else "Child"
        relationships.append({
            "relation": relation,
            "member": member,
        })

    context = {
        "household_id": household_id,
        "head": head,
        "relationships": relationships,
        "total_members": members.count(),
    }
    return render(request, "myproject/household_details.html", context)

# ---------- Deleted Beneficiaries page ---------------------------------------------------------------

@login_required
def deleted_beneficiaries(request):

    is_admin_or_manager = (
        request.user.is_superuser
        or request.user.groups.filter(name__in=["Manager", "Admin"]).exists()
    )

    if is_admin_or_manager:
        submissions = Beneficiary.deleted_objects.all()
    else:
        submissions = Beneficiary.deleted_objects.filter(created_by=request.user)

    deleted_records = submissions.order_by("-deleted_at")

    if request.method == "POST":
        pk = request.POST.get("pk")
        action = request.POST.get("action")

        record = get_object_or_404(Beneficiary.deleted_objects, pk=pk)

        # -------- RESTORE --------
        if action == "restore":
            record.Deleted = False
            record.undeleted_at = timezone.now()
            record.undeleted_by= request.user.username
            record.save()
            messages.success(
                request,
                "✅ The beneficiary has been restored successfully",
                extra_tags="auto",
            )

        # -------- PERMANENT DELETE --------
        elif action == "delete":
            if not is_admin_or_manager:
                messages.error(request, "❌ Permission denied", extra_tags="auto")
                return redirect("deleted_beneficiaries")

            record.delete()
            messages.success(
                request,
                "🗑️ Record permanently deleted",
                extra_tags="auto",
            )

        return redirect("deleted_beneficiaries")

    return render(
        request,
        "myproject/deleted_beneficiaries.html",
        {
            "records": deleted_records,
            "can_delete": is_admin_or_manager,  # ✅ PASS SIMPLE BOOLEAN
        },
    )





# ---------- Adding Beneficiaries page ---------------------------------------------------------------

import uuid
import requests
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Beneficiary
from .households import assign_households


@login_required
def beneficiary_add(request):
    if request.method == "POST":

        user = request.user if request.user.is_authenticated else None

        try:
            # -----------------------------
            # SAVE FIRST (without eligibility)
            # -----------------------------
            beneficiary = Beneficiary.objects.create(
                IP_Name=safe_value(request.POST.get("IP_Name")),
                Sector=safe_value(request.POST.get("Sector")),
                Indicator=safe_value(request.POST.get("Indicator")),
                Date=safe_value(request.POST.get("Date")),
                Name=safe_value(request.POST.get("Name")),
                ID_Number=safe_value(request.POST.get("ID_Number")),
                Parent_ID=safe_value(request.POST.get("Parent_ID")),
                Spouse_ID=safe_value(request.POST.get("Spouse_ID")),
                Phone_Number=safe_value(request.POST.get("Phone_Number")),
                Date_of_Birth=safe_value(request.POST.get("Date_of_Birth")),
                Age=safe_value(request.POST.get("Age")),
                Gender=safe_value(request.POST.get("Gender")),
                Governorate=safe_value(request.POST.get("Governorate")),
                Municipality=safe_value(request.POST.get("Municipality")),
                Neighborhood=safe_value(request.POST.get("Neighborhood")),
                Site_Name=safe_value(request.POST.get("Site_Name")),
                Disability_Status=safe_value(request.POST.get("Disability_Status")),
                Supply_Type=safe_value(request.POST.get("Supply_Type")),
                Benefit_Date=safe_value(request.POST.get("Benefit_Date")),
                HH_Members=safe_value(request.POST.get("HH_Members")),
                Marital_Status=safe_value(request.POST.get("Marital_Status")),
                created_by=user,
            )

            # -----------------------------
            # CALCULATE ELIGIBILITY (CORRECT)
            # -----------------------------
            beneficiary.Eligiblity = calculate_eligibility(beneficiary)
            beneficiary.save(update_fields=["Eligiblity"])

            # -----------------------------
            # HOUSEHOLDS
            # -----------------------------
            changes = assign_households()

            messages.success(
                request,
                f"✅ Beneficiary added successfully. Household updated for {changes} records.",
                extra_tags="auto",
            )

        except Exception as e:
            messages.error(
                request,
                f"❌ Failed to save beneficiary: {str(e)}",
                extra_tags="auto",
            )

        return redirect("home")

    # -----------------------------
    # GET REQUEST (FORM DATA)
    # -----------------------------
    ip_list = (
        APIToken.objects
        .values_list("IP", flat=True)
        .distinct()
        .order_by("IP")
    )

    sector_list = (
        APIToken.objects
        .values_list("section", flat=True)
        .distinct()
        .order_by("section")
    )

    supply_types = (
        Supply.objects
        .exclude(supply_type__isnull=True)
        .exclude(supply_type="")
        .values_list("supply_type", flat=True)
        .distinct()
        .order_by("supply_type")
    )

    context = {
        "ip_list": ip_list,
        "sector_list": sector_list,
        "supply_types": supply_types,
    }

    return render(request, "myproject/beneficiary_add.html", context)


# ---------- Manually run Assiogn_Households ---------------------------------------------------------------

@login_required
@role_required('Manager', 'Admin')
def assign_households_view(request):
    try:
        # Run your existing Python function
        result = assign_households()
        return JsonResponse({"status": "success", "message": "Households reassigned successfully."})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})
    

# ---------- SERIALIZERs ---------------------------------------------------------------


from rest_framework import viewsets, permissions
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from django.utils import timezone
from rest_framework_simplejwt.authentication import JWTAuthentication

from .models import Beneficiary
from .serializers import BeneficiarySerializer


# Pagination: 100 per page
class BeneficiaryPagination(PageNumberPagination):
    page_size = 100
    max_page_size = 1000


class BeneficiaryViewSet(viewsets.ModelViewSet):
    serializer_class = BeneficiarySerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = BeneficiaryPagination

    # -----------------------------------------
    # FILTERING BY USER ROLE
    # -----------------------------------------
    
    
# -----------------------------------------
    # FILTERING BY USER ROLE + updated_after
    # -----------------------------------------
    def get_queryset(self):
        user = self.request.user

        # Base queryset: include deleted records via all_objects
        if user.is_superuser or user.groups.filter(name__in=["Admin", "Manager"]).exists():
            qs = Beneficiary.all_objects.all().order_by("-created_at")
        else:
            qs = Beneficiary.all_objects.filter(created_by=user).order_by("-created_at")

        # Apply updated_after filter if present (no timezone manipulation)
        updated_after = self.request.query_params.get("updated_after")
        if updated_after:
            dt = parse_datetime(updated_after)
            if dt:
                qs = qs.filter(updated_at__gt=dt)

        return qs

    # -----------------------------------------
    # LIST: preserve pagination, append server_time
    # -----------------------------------------
    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        data = response.data

        # data is an OrderedDict from pagination with keys: count, next, previous, results
        payload = {
            "count": data.get("count"),
            "next": data.get("next"),
            "previous": data.get("previous"),
            "results": data.get("results"),
            "server_time": timezone.now().isoformat(),
        }
        return Response(payload, status=response.status_code)

    # -----------------------------------------
    # RETRIEVE: add server_time to single-item response
    # -----------------------------------------
    def retrieve(self, request, *args, **kwargs):
        response = super().retrieve(request, *args, **kwargs)
        payload = {
            "server_time": timezone.now().isoformat(),
            "result": response.data,
               }
        
        return Response(payload, status=response.status_code)


    # -----------------------------------------
    # CREATE
    # -----------------------------------------
    def perform_create(self, serializer):
        serializer.save(
            created_by=self.request.user,
            Submission_Time=timezone.now()
        )
        

    # -----------------------------------------
    # UPDATE (partial update allowed)
    # -----------------------------------------
    def update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return super().update(request, *args, **kwargs)




#===============================BULK IMPORT ===================================
# 
# views.py (add after your other views)

# Global progress for bulk import (separate from sync_data)
bulk_progress_status = {
    "stage": "Fetching Data",
    "progress": 0,
    "total": 0,
    "processed": 0,
    "created": 0,
    "failed": 0,
    "errors": [],
    "completed": False,
}

@login_required
def bulk_upload_page(request):
    """
    Render the page where user uploads Excel and picks IP.
    """
    ip_list = APIToken.objects.values_list('IP', flat=True).distinct().order_by('IP')
    context = {
        "ip_list": ip_list,
    }
    return render(request, "myproject/bulk_upload.html", context)


import os
import uuid
import threading
import tempfile
from openpyxl import load_workbook

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.conf import settings

from .models import Beneficiary  # Adjust import as needed


# Global progress tracker (in production, consider Redis or DB for multi-worker safety)
bulk_progress_status = {
    "stage": "Idle",
    "progress": 0,
    "total": 0,
    "processed": 0,
    "created": 0,
    "failed": 0,
    "errors": [],
    "completed": False,
}


@login_required
def start_bulk_import(request):
    """
    Start background thread to parse uploaded Excel file and import beneficiaries
    directly into the local database (no InForm API interaction).
    Expects POST with 'excel_file' and 'ip_name'.
    """
    global bulk_progress_status

    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Only POST allowed."}, status=400)

    excel_file = request.FILES.get("excel_file")
    ip_name = request.POST.get("ip_name")

    if not excel_file:
        return JsonResponse({"status": "error", "message": "No file uploaded."}, status=400)
    if not ip_name:
        return JsonResponse({"status": "error", "message": "Please choose an Implementing Partner (IP)."}, status=400)

    # Save uploaded file temporarily
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        for chunk in excel_file.chunks():
            tmp.write(chunk)
        tmp.flush()
        tmp.close()
    except Exception as e:
        os.unlink(tmp.name)
        return JsonResponse({"status": "error", "message": f"Failed saving file: {str(e)}"}, status=500)

    # Reset progress
    bulk_progress_status = {
        "stage": "Queued",
        "progress": 0,
        "total": 0,
        "processed": 0,
        "created": 0,
        "failed": 0,
        "errors": [],
        "completed": False,
    }

    def run_import(file_path, ip_name):
        global bulk_progress_status
        user = request.user  # Capture user from request context

        try:
            bulk_progress_status.update({"stage": "Reading Excel file", "progress": 5})

            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            sheet = wb.active

            rows = sheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                bulk_progress_status.update({
                    "stage": "❌ Excel file is empty",
                    "progress": 100,
                    "completed": True,
                    "errors": ["Excel file contains no data rows."]
                })
                return

            # Normalize header
            header_map = [str(col).strip() if col is not None else f"col_{i}" for i, col in enumerate(header)]

            # Field mapping: Excel column → Beneficiary field (case-insensitive + aliases)
            field_aliases = {
                "full name": "Name", "name": "Name",
                "id": "ID_Number", "id number": "ID_Number", "national id": "ID_Number",
                "phone": "Phone_Number", "phone number": "Phone_Number",
                "dob": "Date_of_Birth", "date of birth": "Date_of_Birth",
                "age": "Age", "gender": "Gender",
                "governorate": "Governorate", "municipality": "Municipality",
                "neighborhood": "Neighborhood", "site": "Site_Name", "site name": "Site_Name",
                "disability": "Disability_Status", "disability status": "Disability_Status",
                "sector": "Sector", "indicator": "Indicator", "date": "Date",
                "ip": "IP_Name", "ip name": "IP_Name", "ip_name": "IP_Name",
            }

            expected_fields = [
                "IP_Name", "Sector", "Indicator", "Date", "Name", "ID_Number",
                "Parent_ID", "Spouse_ID", "Phone_Number", "Date_of_Birth", "Age",
                "Gender", "Governorate", "Municipality", "Neighborhood",
                "Site_Name", "Disability_Status"
            ]

            # Resolve mapping: header → standard field name
            col_to_field = {}
            for idx, col_name in enumerate(header_map):
                lower_col = col_name.lower()
                mapped = next((f for f in expected_fields if f.lower() == lower_col), None)
                if not mapped:
                    mapped = field_aliases.get(lower_col)
                col_to_field[idx] = mapped or col_name  # fallback to original if no match

            # Parse data rows
            records = []
            for row in rows:
                if not any(cell is not None and str(cell).strip() != "" for cell in row):
                    continue  # skip empty rows
                rec = {}
                for idx, value in enumerate(row):
                    field_name = col_to_field.get(idx)
                    if field_name:
                        rec[field_name] = None if value is None else str(value).strip()
                # Ensure IP_Name is set
                rec["IP_Name"] = rec.get("IP_Name") or ip_name
                records.append(rec)

            total = len(records)
            bulk_progress_status.update({
                "stage": "Processing records",
                "total": total,
                "progress": 10 if total > 0 else 100
            })

            if total == 0:
                bulk_progress_status.update({
                    "stage": "❌ No valid data rows found",
                    "completed": True,
                    "errors": ["No valid data rows in Excel."]
                })
                return

            created = 0
            failed = 0
            errors = []

            for idx, rec in enumerate(records, start=1):
                try:
                    id_number = rec.get("ID_Number")
                    if not id_number:
                        errors.append(f"Row {idx}: Missing ID_Number. Skipped.")
                        failed += 1
                        continue

                    # Create Beneficiary directly
                    Beneficiary.all_objects.create(
                        IP_Name=rec.get("IP_Name"),
                        Sector=rec.get("Sector"),
                        Indicator=rec.get("Indicator"),
                        Date=rec.get("Date"),
                        Name=rec.get("Name"),
                        ID_Number=id_number,
                        Parent_ID=rec.get("Parent_ID"),
                        Spouse_ID=rec.get("Spouse_ID"),
                        Phone_Number=rec.get("Phone_Number"),
                        Date_of_Birth=rec.get("Date_of_Birth"),
                        Age=rec.get("Age"),
                        Gender=rec.get("Gender"),
                        Governorate=rec.get("Governorate"),
                        Municipality=rec.get("Municipality"),
                        Neighborhood=rec.get("Neighborhood"),
                        Site_Name=rec.get("Site_Name"),
                        Disability_Status=rec.get("Disability_Status"),
                        created_by=user,
                        # Optional: set Submission_Time to now
                        created_at=timezone.now() if hasattr(Beneficiary, 'Submission_Time') else None,
                    )

                    created += 1

                except Exception as e:
                    failed += 1
                    errors.append(f"Row {idx}: {str(e)[:150]}")

                # Update progress
                progress = 10 + int((idx / total) * 80)  # 10%–90%
                bulk_progress_status.update({
                    "stage": f"Processing row {idx}/{total}",
                    "progress": progress,
                    "processed": idx,
                    "created": created,
                    "failed": failed,
                    "errors": errors[:10],  # limit visible errors
                })

            # Final steps: household assignment
            bulk_progress_status.update({"stage": "Assigning households...", "progress": 92})
            try:
                changes = assign_households()
                errors.append(f"Household assignment completed: {changes} records updated.")
            except Exception as e:
                errors.append(f"Household assignment failed: {str(e)[:150]}")

            # Complete
            bulk_progress_status.update({
                "stage": "✅ Import completed successfully",
                "progress": 100,
                "processed": total,
                "created": created,
                "failed": failed,
                "errors": errors,
                "completed": True,
            })

        except Exception as e:
            bulk_progress_status.update({
                "stage": f"❌ Fatal error: {str(e)[:200]}",
                "progress": 100,
                "completed": True,
                "errors": [str(e)],
            })
        finally:
            # Clean up temp file
            try:
                os.unlink(file_path)
            except OSError:
                pass

    # Start background thread
    thread = threading.Thread(target=run_import, args=(tmp.name, ip_name), daemon=True)
    thread.start()

    return JsonResponse({"status": "started", "message": "Bulk import started in background."})


@login_required
def get_bulk_progress(request):
    """
    Return current bulk import progress as JSON.
    """
    global bulk_progress_status
    return JsonResponse(bulk_progress_status)


def autocomplete(request):
    field = request.GET.get("field")
    query = request.GET.get("q", "")

    if not field:
        return JsonResponse([], safe=False)

    # Get unique values for that field
    values = (
        Beneficiary.objects
        .filter(**{f"{field}__icontains": query})
        .values_list(field, flat=True)
        .distinct()
    )

    return JsonResponse(list(values[:15]), safe=False)
    


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_user_info(request):
    user = request.user
    return Response({
        'username': user.username,
        'firstname': user.first_name,
        'lastname': user.last_name,
        "full_name": f"{user.first_name} {user.last_name}".strip(),
        'email': user.email,
        'groups': list(user.groups.values_list('name', flat=True))
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_duplicate(request, id_number):
    exists = Beneficiary.objects.filter(id_number=id_number).exists()
    return Response({"exists": exists})


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import Beneficiary


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def distinct_ip_names(request):
    ip_names = (
        APIToken.objects
        .exclude(IP_Name__isnull=True)
        .exclude(IP_Name__exact='')
        .values_list('IP_Name', flat=True)
        .distinct()
        .order_by('IP_Name')
    )
    return Response({"ip_names": list(ip_names)})



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def distinct_sectors(request):
    sectors = (
        APIToken.objects
        .exclude(Section__isnull=True)
        .exclude(Section__exact='')
        .values_list('Section', flat=True)
        .distinct()
        .order_by('Section')
    )
    return Response({"sectors": list(sectors)})

#====================================================================================
# #=============================== SUPPLY SECTION ===================================
# #==================================================================================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Supply

@login_required
def supply_list(request):
    supplies = Supply.objects.all().order_by('-created_at')
    return render(request, 'myproject/supply_list.html', {'supplies': supplies})

@login_required
def supply_add(request):
    ip_list = (
        APIToken.objects
        .exclude(IP__isnull=True)
        .exclude(IP='')
        .values_list('IP', flat=True)
        .distinct()
        .order_by('IP')
    )

    section_list = (
        APIToken.objects
        .exclude(section__isnull=True)
        .exclude(section='')
        .values_list('section', flat=True)
        .distinct()
        .order_by('section')
    )

    if request.method == 'POST':
        Supply.objects.create(
            section=request.POST.get('section'),
            ip_name=request.POST.get('ip_name'),
            supply_type=request.POST.get('supply_type'),
            eligibility_period=request.POST.get('eligibility_period'),
            distribution_date=request.POST.get('distribution_date') or None
        )
        messages.success(request, "Supply added successfully")
        return redirect('supply_list')

    return render(request, 'myproject/supply_add.html', {
        'ip_list': ip_list,
        'section_list': section_list,
    })


@login_required
def supply_update(request, pk):
    supply = get_object_or_404(Supply, pk=pk)

    ip_list = (
        APIToken.objects
        .exclude(IP__isnull=True)
        .exclude(IP='')
        .values_list('IP', flat=True)
        .distinct()
        .order_by('IP')
    )

    section_list = (
        APIToken.objects
        .exclude(section__isnull=True)
        .exclude(section='')
        .values_list('section', flat=True)
        .distinct()
        .order_by('section')
    )

    if request.method == 'POST':
        supply.section = request.POST.get('section')
        supply.ip_name = request.POST.get('ip_name')
        supply.eligibility_period = request.POST.get('eligibility_period')
        supply.distribution_date = request.POST.get('distribution_date')
        supply.save()

        messages.success(request, "Supply updated successfully")
        return redirect('supply_list')

    return render(request, 'myproject/supply_update.html', {
        'supply': supply,
        'ip_list': ip_list,
        'section_list': section_list,
    })

@login_required
def supply_delete(request, pk):
    supply = get_object_or_404(Supply, pk=pk)
    supply.delete()
    return JsonResponse({'success': True})

#===============================================================================================


def calculate_eligibility(beneficiary):
    """
    Returns 'True' or 'False' based on supply eligibility rules
    """
    #print(start_date)
    #print(beneficiary)
    
    if  beneficiary is None:
        return True
    
    if not beneficiary.Supply_Type:
        return True
    
    if not beneficiary.Benefit_Date:
        return True

    # Get eligibility period from Supply
    supply = (
        Supply.objects
        .filter(supply_type=beneficiary.Supply_Type)
        .order_by('-created_at')
        .first()
    )

    if not supply or not supply.eligibility_period:
        return True

    try:
        eligibility_days = int(supply.eligibility_period)
    except ValueError:
        return False

    start_date = beneficiary.Benefit_Date - timedelta(days=eligibility_days)

    # Check for previous benefits with same supply
    

    qs = Beneficiary.all_objects.filter(
        ID_Number=beneficiary.ID_Number,
        Supply_Type=beneficiary.Supply_Type,
        
    )

    # ✅ Exclude current record ONLY if it exists in DB
    if hasattr(beneficiary, "pk") and beneficiary.pk:
        qs = qs.exclude(pk=beneficiary.pk)

    previous = qs.exists()

    #print(start_date)
    #print(beneficiary.Supply_Type)
    #print(beneficiary.Benefit_Date)
    #print(previous)
    #print(beneficiary.ID_Number)


    return False if previous else True


#===============SUPPLY DASHBOARD+++++++++++++++++++++++++++++++++++
from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Max
from .models import Supply, Beneficiary, APIToken


# =========================================================
# SUPPLY DASHBOARD (KPIs + Charts + Filters)
# =========================================================

@login_required
def supply_dashboard(request):
    # ---------------- FILTER VALUES ----------------
    selected_ip = request.GET.get('ip')
    selected_section = request.GET.get('section')

    ip_list = (
        APIToken.objects
        .exclude(IP__isnull=True)
        .exclude(IP='')
        .values_list('IP', flat=True)
        .distinct()
        .order_by('IP')
    )

    section_list = (
        APIToken.objects
        .exclude(section__isnull=True)
        .exclude(section='')
        .values_list('section', flat=True)
        .distinct()
        .order_by('section')
    )

    # ---------------- BASE BENEFICIARY QUERY ----------------
    beneficiaries = Beneficiary.objects.exclude(
        Supply_Type__isnull=True
    ).exclude(
        Supply_Type=''
    )

    if selected_ip:
        beneficiaries = beneficiaries.filter(IP_Name=selected_ip)

    if selected_section:
        beneficiaries = beneficiaries.filter(Sector=selected_section)

    # ---------------- KPI COUNTS ----------------
    total_supplies = Supply.objects.count()
    total_benefits = beneficiaries.count()
    eligible_count = beneficiaries.filter(Eligiblity=True).count()
    not_eligible_count = beneficiaries.filter(Eligiblity=False).count()

    # ---------------- TABLE + CHART DATA ----------------
    usage_data = []
    chart_labels = []
    chart_total = []
    chart_eligible = []
    chart_not_eligible = []

    supplies = Supply.objects.all().order_by('supply_type')

    for supply in supplies:
        qs = beneficiaries.filter(Supply_Type=supply.supply_type)
        total = qs.count()

        if total == 0:
            continue

        eligible = qs.filter(Eligiblity=True).count()
        not_eligible = qs.filter(Eligiblity=False).count()

        usage_data.append({
            "supply_type": supply.supply_type,
            "eligibility_period": supply.eligibility_period,
            "total": total,
            "eligible": eligible,
            "not_eligible": not_eligible,
            "last_date": qs.aggregate(Max('Benefit_Date'))['Benefit_Date__max'],
        })

        chart_labels.append(supply.supply_type)
        chart_total.append(total)
        chart_eligible.append(eligible)
        chart_not_eligible.append(not_eligible)

    return render(request, "myproject/supply_dashboard.html", {
        "ip_list": ip_list,
        "section_list": section_list,
        "selected_ip": selected_ip,
        "selected_section": selected_section,

        "total_supplies": total_supplies,
        "total_benefits": total_benefits,
        "eligible_count": eligible_count,
        "not_eligible_count": not_eligible_count,

        "usage_data": usage_data,
        "chart_labels": chart_labels,
        "chart_total": chart_total,
        "chart_eligible": chart_eligible,
        "chart_not_eligible": chart_not_eligible,
    })


# =========================================================
# SUPPLY → BENEFICIARIES (DRILL-DOWN)
# =========================================================
from django.core.paginator import Paginator

@login_required
def supply_beneficiaries(request, supply_type):
    selected_ip = request.GET.get("ip")
    selected_section = request.GET.get("section")

    # ✅ Normalize bad values coming from querystring
    if selected_ip in ("", "None", None):
        selected_ip = None

    if selected_section in ("", "None", None):
        selected_section = None

    # ✅ CORRECT FIELD NAME
    beneficiaries = Beneficiary.objects.filter(
        Supply_Type=supply_type
    )

    if selected_ip:
        beneficiaries = beneficiaries.filter(IP_Name=selected_ip)

    if selected_section:
        beneficiaries = beneficiaries.filter(Sector=selected_section)

    beneficiaries = beneficiaries.order_by("-Benefit_Date")

    paginator = Paginator(beneficiaries, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "myproject/supply_beneficiaries.html", {
        "supply_type": supply_type,
        "page_obj": page_obj,
        "selected_ip": selected_ip,
        "selected_section": selected_section,
    })



#+++++++++++++++++++++++++EXCEL EXPORT FOR SUPPLY++++++++++++++++++++++++++++++++++

@login_required
def export_supply_dashboard_excel(request):
    if not request.user.groups.filter(name__in=["Admin", "Manager"]).exists():
        return HttpResponseForbidden("Not authorized")

    selected_ip = request.GET.get('ip')
    selected_section = request.GET.get('section')

    beneficiaries = Beneficiary.objects.exclude(
        Supply_Type__isnull=True
    ).exclude(
        Supply_Type=''
    )

    if selected_ip:
        beneficiaries = beneficiaries.filter(IP_Name=selected_ip)

    if selected_section:
        beneficiaries = beneficiaries.filter(Sector=selected_section)

    wb = Workbook()
    ws = wb.active
    ws.title = "Supply Summary"

    headers = [
        "Supply Type",
        "Eligibility Period (Days)",
        "Total Beneficiaries",
        "Eligible",
        "Not Eligible",
        "Last Benefit Date",
        "Filtered IP",
        "Filtered Section",
        "Exported At",
        "Exported By",
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for supply in Supply.objects.all().order_by('supply_type'):
        qs = Beneficiary.objects.filter(
            Supply_Type=supply.supply_type
        ).exclude(
            ID_Number__isnull=True
        )


        ws.append([
            supply.supply_type,
            supply.eligibility_period,
            qs.count(),
            qs.filter(Eligiblity=True).count(),
            qs.filter(Eligiblity=False).count(),
            qs.aggregate(Max('Benefit_Date'))['Benefit_Date__max'],
            selected_ip or "All",
            selected_section or "All",
            timezone.now().strftime("%Y-%m-%d %H:%M"),
            request.user.get_full_name() or request.user.username,
        ])



    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"supply_dashboard_summary_{timezone.now().date()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response

def excel_datetime(dt):
    if not dt:
        return None
    if hasattr(dt, "tzinfo") and dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt


@login_required
def export_supply_beneficiaries_excel(request, supply_type):
    if not request.user.groups.filter(name__in=["Admin", "Manager"]).exists():
        return HttpResponseForbidden("Not authorized")

    selected_ip = request.GET.get("ip")
    selected_section = request.GET.get("section")

    # ✅ NORMALIZE QUERY PARAMS
    if selected_ip in ("", "None", None):
        selected_ip = None

    if selected_section in ("", "None", None):
        selected_section = None

    qs = Beneficiary.objects.filter(
        Supply_Type=supply_type
        ).exclude(
            ID_Number__isnull=True
        )

    if selected_ip:
        qs = qs.filter(IP_Name=selected_ip)

    if selected_section:
        qs = qs.filter(Sector=selected_section)

    wb = Workbook()
    ws = wb.active
    ws.title = "Beneficiaries"

    headers = [
        "Name",
        "ID Number",
        "IP Name",
        "Sector",
        "Supply Type",
        "Benefit Date",
        "Eligibility",
        "Eligibility Period (Days)",
        "Eligibility Expiry Date",
        "Household ID",
        "Submission Time",
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    supply = Supply.objects.filter(supply_type=supply_type).first()
    eligibility_days = int(supply.eligibility_period) if supply else 0

    for b in qs:
        expiry_date = None
        if b.Benefit_Date and eligibility_days:
            expiry_date = b.Benefit_Date + timedelta(days=eligibility_days)

        ws.append([
            b.Name,
            b.ID_Number,
            b.IP_Name,
            b.Sector,
            b.Supply_Type,
            excel_datetime(b.Benefit_Date),
            "Eligible" if b.Eligiblity else "Not Eligible",
            eligibility_days,
            excel_datetime(expiry_date),
            b.Household_ID,
            excel_datetime(b.Submission_Time),
        ])


    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"supply_{supply_type}_beneficiaries_{timezone.now().date()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


#=============================BULK SUPPLY IMPORT++++===============================

import io
import pandas as pd
from django.db import transaction
from django.utils.timezone import now

# ================= BULK SUPPLY =================

@login_required
def bulk_supply_bulk_page(request):
    request.session.pop("bulk_supply_preview", None)

    return render(request, "myproject/bulk_supply_import.html", {
        "section_list": APIToken.objects.values_list("section", flat=True).distinct().order_by("section"),
        "ip_list": APIToken.objects.values_list("IP", flat=True).distinct().order_by("IP"),
        "supply_list": Supply.objects.values_list("supply_type", flat=True).distinct()
    })


import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.utils.timezone import now

from .models import APIToken, Supply, Beneficiary  # adjust import if needed


# ==============================
# 1) IMPORT PAGE (UPLOAD FORM)
# ==============================
@login_required
def bulk_supply_import_page(request):
    """
    GET  -> show upload page
    POST -> validate, read excel, build preview rows in session, then redirect to review page
    """
    section_list = APIToken.objects.values_list("section", flat=True).distinct().order_by("section")
    ip_list = APIToken.objects.values_list("IP", flat=True).distinct().order_by("IP")
    supply_list = Supply.objects.values_list("supply_type", flat=True).distinct().order_by("supply_type")

    if request.method == "GET":
        # Optional: clear stale preview when opening page
        request.session.pop("bulk_supply_preview", None)
        request.session.pop("bulk_supply_results", None)

        return render(request, "myproject/bulk_supply_import.html", {
            "section_list": section_list,
            "ip_list": ip_list,
            "supply_list": supply_list,
        })

    # ---------- POST ----------
    excel = request.FILES.get("excel_file")
    section = (request.POST.get("section") or "").strip() or None
    ip_name = (request.POST.get("ip_name") or "").strip() or None
    supply_type = (request.POST.get("supply_type") or "").strip()
    benefit_date = (request.POST.get("benefit_date") or "").strip()

    # Professional messages
    if not section:
        messages.error(request, "Please select a Section before proceeding.")
        return redirect("bulk_supply_import")
    
    if not ip_name:
        messages.error(request, "Please select an IP Name before proceeding.")
        return redirect("bulk_supply_import")
    
    if not supply_type:
        messages.error(request, "Please select a Supply Type before proceeding.")
        return redirect("bulk_supply_import")

    if not benefit_date:
        messages.error(request, "Please select the Benefit Date before proceeding.")
        return redirect("bulk_supply_import")
    
    if not excel:
        messages.error(request, "Please upload the Excel file using the provided template.")
        return redirect("bulk_supply_import")
    
    

    # Parse benefit date
    try:
        benefit_dt = datetime.strptime(benefit_date, "%Y-%m-%d").date()
    except ValueError:
        messages.error(request, "Benefit Date is invalid. Please select a valid date.")
        return redirect("bulk_supply_import")

    # Read excel
    try:
        df = pd.read_excel(excel)
    except Exception:
        messages.error(request, "The uploaded file could not be read. Please upload a valid Excel (.xlsx) file.")
        return redirect("bulk_supply_import")

    # Required columns
    required_cols = {"ID_Number", "Name"}
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        messages.error(
            request,
            f"Missing required column(s): {', '.join(missing)}. Please use the official template."
        )
        return redirect("bulk_supply_import")

    # Load supply & period
    supply = Supply.objects.filter(supply_type=supply_type).first()
    if not supply:
        messages.error(request, "Selected Supply Type was not found. Please check the Supply setup.")
        return redirect("bulk_supply_import")

    try:
        period_days = int(supply.eligibility_period)
    except Exception:
        messages.error(request, "Eligibility period is not configured correctly for this Supply Type.")
        return redirect("bulk_supply_import")

    # Build preview rows
    rows = []
    for _, r in df.iterrows():
        id_number = str(r.get("ID_Number") or "").strip()
        name = str(r.get("Name") or "").strip()
        
        if not id_number or not name:
            # skip empty / invalid rows silently (or you can collect errors)
            continue

        last = Beneficiary.objects.filter(
            ID_Number=id_number,
            Supply_Type=supply_type
        ).order_by("-Benefit_Date").first()

        eligible = True
        if last and last.Benefit_Date:
            try:
                delta_days = (benefit_dt - last.Benefit_Date).days
                if delta_days < period_days:
                    eligible = False
            except Exception:
                # If anything weird in dates, default to not eligible (safer)
                eligible = False

        rows.append({
            "name": name,
            "id_number": id_number,
            "section": section,
            "ip_name": ip_name,
            "supply": supply_type,
            "benefit_date": benefit_date,  # keep as string for session safety
            "eligibility": eligible,        # bool
            "override": False,
            "override_by": None,
            "override_at": None,
        })

    if not rows:
        messages.warning(request, "No valid records were found in the Excel sheet.")
        return redirect("bulk_supply_import")

    request.session["bulk_supply_preview"] = rows
    request.session["bulk_supply_meta"] = {
        "ip_name": ip_name,
        "supply_type": supply_type,
        "benefit_date": benefit_date,
        "total_rows": len(rows),
    }

    return redirect("bulk_supply_review")


# ==============================
# 2) REVIEW PAGE (SHOW PREVIEW)
# ==============================
@login_required
def bulk_supply_review(request):
    """
    Renders the review page from session. No excel reading here.
    """
    rows = request.session.get("bulk_supply_preview", [])
    meta = request.session.get("bulk_supply_meta", {})

    if not rows:
        messages.warning(request, "There is no preview data to review. Please upload the Excel file again.")
        return redirect("bulk_supply_import")

    return render(request, "myproject/bulk_supply_review.html", {
        "rows": rows,
        "meta": meta,
    })



from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from datetime import datetime
import pandas as pd


@login_required
def bulk_supply_preview(request):
    if request.method != "POST":
        return redirect("bulk_supply_import")

    excel = request.FILES.get["excel_file"]
    supply_type = request.POST.get["supply_type"]
    benefit_date = request.POST.get["benefit_date"]
    ip_name = request.POST.get("ip_name")
    section = request.POST.get["section"]

    # Professional messages
    if not excel:
        messages.error(request, "Please upload the Excel file using the provided template.")
        return redirect("bulk_supply_import")
    
    if not section:
        messages.error(request, "Please select a Section before proceeding.")
        return redirect("bulk_supply_import")
    
    if not ip_name:
        messages.error(request, "Please select an IP Name before proceeding.")
        return redirect("bulk_supply_import")
    
    if not supply_type:
        messages.error(request, "Please select a Supply Type before proceeding.")
        return redirect("bulk_supply_import")

    if not benefit_date:
        messages.error(request, "Please select the Benefit Date before proceeding.")
        return redirect("bulk_supply_import")

    df = pd.read_excel(excel)
    supply = Supply.objects.get(supply_type=supply_type)
    benefit_dt = datetime.strptime(benefit_date, "%Y-%m-%d").date()

    rows = []
    for _, r in df.iterrows():
        last = Beneficiary.objects.filter(
            ID_Number=str(r["ID_Number"]),
            Supply_Type=supply_type
        ).order_by("-Benefit_Date").first()

        eligible = True
        if last and last.Benefit_Date:
            delta = (benefit_dt - last.Benefit_Date).days
            if delta < int(supply.eligibility_period):
                eligible = False

        rows.append({
            "name": r["Name"],
            "id_number": str(r["ID_Number"]),
            "section": section,
            "ip_name": ip_name,
            "supply": supply_type,
            "benefit_date": benefit_date,
            "eligibility": eligible,
            "override": False,
        })

    request.session["bulk_supply_preview"] = rows
    return redirect("bulk_supply_review")




from django.utils.timezone import now

@login_required
def bulk_supply_override_eligibility(request):
    idx = int(request.POST["index"])
    preview = request.session["bulk_supply_preview"]

    preview[idx]["eligibility"] = True
    preview[idx]["override"] = True
    preview[idx]["override_by"] = request.user.username
    preview[idx]["override_at"] = now().isoformat()

    request.session["bulk_supply_preview"] = preview
    return redirect("bulk_supply_review")


@login_required
def bulk_supply_delete_preview_row(request):
    idx = int(request.POST["index"])
    preview = request.session["bulk_supply_preview"]
    preview.pop(idx)
    request.session["bulk_supply_preview"] = preview
    return redirect("bulk_supply_review")




from django.db import transaction
from django.http import JsonResponse

from django.db import transaction
from django.shortcuts import redirect

import uuid

@login_required
def bulk_supply_import_commit(request):
    preview = request.session.get("bulk_supply_preview")

    if not preview:
        messages.error(request, "Nothing to submit.")
        return redirect("bulk_supply_import")

    batch_id = uuid.uuid4()

    with transaction.atomic():
        for r in preview:
            Beneficiary.objects.create(
                Name=r["name"],
                ID_Number=r["id_number"],
                IP_Name=r["ip_name"],
                Sector=r["section"],
                Supply_Type=r["supply"],
                Benefit_Date=r["benefit_date"],
                Eligiblity=r["eligibility"],
                supply_batch_id=batch_id,
                created_by=request.user.username
            )

    request.session.pop("bulk_supply_preview", None)

    return redirect("bulk_supply_done", batch_id=batch_id)



import io
from openpyxl import Workbook
from django.http import HttpResponse

@login_required
def bulk_supply_export_excel(request, batch_id):
    qs = Beneficiary.objects.filter(supply_batch_id=batch_id)

    if not qs.exists():
        return HttpResponse("No data found for this batch.", status=400)

    wb = Workbook()

    def write(ws, items):
        ws.append([
            "Name", "ID_Number", "Sector", "IP_Name", "Supply_Type",
            "Benefit_Date", "Eligibility"
        ])
        for b in items:
            ws.append([
                b.Name,
                b.ID_Number,
                b.Sector,
                b.IP_Name,
                b.Supply_Type,
                str(b.Benefit_Date),
                "Eligible" if b.Eligiblity else "Not Eligible"
            ])

    ws1 = wb.active
    ws1.title = "Eligible"
    write(ws1, qs.filter(Eligiblity=True))

    ws2 = wb.create_sheet("Not Eligible")
    write(ws2, qs.filter(Eligiblity=False))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=bulk_supply_{batch_id}.xlsx"
    return response



@login_required
def bulk_supply_done(request, batch_id):
    return render(
        request,
        "myproject/bulk_supply_done.html",
        {"batch_id": batch_id}
    )



#============================SUPPLY Check =====================================

# views.py

@login_required
def eligibility_check(request):
    return render(request, "myproject/eligibility_check.html", {
        "ip_list": APIToken.objects.values_list("IP", flat=True).distinct().order_by("IP"),
        "sector_list": APIToken.objects.values_list("section", flat=True).distinct().order_by("section"),
        "supply_list": Supply.objects.values_list("supply_type", flat=True).distinct().order_by("supply_type"),
        "today": timezone.now().date(),
    })


from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from datetime import datetime
from .models import Beneficiary, Supply

@login_required
def eligibility_evaluate(request):
    beneficiary_id = request.GET.get("beneficiary_id")
    supply_type = request.GET.get("supply_type")
    benefit_date = request.GET.get("benefit_date")

    if not (beneficiary_id and supply_type and benefit_date):
        return JsonResponse({"error": "Missing parameters"}, status=400)

    base = Beneficiary.objects.get(id=beneficiary_id)

    benefit_date = datetime.strptime(benefit_date, "%Y-%m-%d").date()

    # 🔹 Last record for same ID & supply
    last = Beneficiary.objects.filter(
        ID_Number=base.ID_Number,
        Supply_Type=supply_type
    ).order_by("-Benefit_Date").first()

    # 🔹 Create TEMP beneficiary object (NOT SAVED)
    temp = Beneficiary(
        ID_Number=base.ID_Number,
        Supply_Type=supply_type,
        Benefit_Date=benefit_date,
        IP_Name=base.IP_Name,
        Sector=base.Sector,
        Date_of_Birth=base.Date_of_Birth,
        Age=base.Age,
        Gender=base.Gender,
        Marital_Status=base.Marital_Status,
    )

    qs = Beneficiary.all_objects.filter(
        ID_Number=temp.ID_Number,
        Supply_Type=temp.Supply_Type,
        
    )

    eligible = not qs.exists()

    #eligible = calculate_eligibility(last or temp)

    return JsonResponse({
        "eligible": eligible,
        "last_benefit_date": last.Benefit_Date if last else None,
    })



@login_required
def eligibility_search(request):
    q = (request.GET.get("q") or "").strip()

    if len(q) < 3:
        return JsonResponse({"results": []})

    qs = (
        Beneficiary.objects
        .filter(Q(ID_Number__icontains=q) | Q(Name__icontains=q))
        .order_by("-Benefit_Date")
    )

    data = []
    for b in qs:
        data.append({
            "id": b.id,
            "name": b.Name,
            "id_number": b.ID_Number,
            "ip_name": b.IP_Name,
            "sector": b.Sector,
            "supply_type": b.Supply_Type,
            "benefit_date": b.Benefit_Date,
        })

    return JsonResponse({"results": data})


from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from datetime import datetime
import uuid
import traceback

@login_required
@require_POST
def eligibility_submit(request):
    try:
        # ---------------- INPUT ----------------
        sector = request.POST.get("sector")
        ip_name = request.POST.get("ip_name")
        beneficiary_id = request.POST.get("beneficiary_id")
        supply_type = request.POST.get("supply_type")
        benefit_date = request.POST.get("benefit_date")
        override = request.POST.get("override") == "1"

        if not beneficiary_id or not supply_type or not benefit_date or not sector or not ip_name:
            return JsonResponse(
                {"warning": "Missing required data."},
                status=400
            )

        benefit_date = datetime.strptime(benefit_date, "%Y-%m-%d").date()

        base = Beneficiary.objects.filter(pk=beneficiary_id).first()
        if not base:
            return JsonResponse(
                {"error": "Beneficiary not found."},
                status=404
            )

        # ---------------- TEMP OBJECT ----------------
        class TempBeneficiary:
            ID_Number = base.ID_Number
            Supply_Type = supply_type
            Benefit_Date = benefit_date
            pk = None  # IMPORTANT

        eligible = calculate_eligibility(TempBeneficiary)

        

        # ---------------- CREATE RECORD ----------------
        with transaction.atomic():
            Beneficiary.objects.create(
                Name=base.Name,
                ID_Number=base.ID_Number,
                IP_Name=base.IP_Name,
                Sector=base.Sector,
                Phone_Number=base.Phone_Number,
                Date_of_Birth=base.Date_of_Birth,
                Age=base.Age,
                Gender=base.Gender,
                Marital_Status=base.Marital_Status,
                Governorate=base.Governorate,
                Municipality=base.Municipality,
                Neighborhood=base.Neighborhood,
                Site_Name=base.Site_Name,
                Disability_Status=base.Disability_Status,
                HH_Members=base.HH_Members,

                Supply_Type=supply_type,
                Benefit_Date=benefit_date,
                Eligiblity=True,

                supply_batch_id=uuid.uuid4(),
                created_by=request.user.username,
            )

        # ✅ ALWAYS RETURN JSON
        return JsonResponse({
            "status": "ok",
            "eligible": eligible,
            "message": "Supply submitted successfully."
        })

    except Exception as e:
        traceback.print_exc()
        return JsonResponse({
            "error": "Server error occurred.",
            "details": str(e)
        }, status=500)


